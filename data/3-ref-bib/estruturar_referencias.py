"""Lê os arquivos _refs.txt extraídos e estrutura cada referência
em campos (autor, título, periódico, ano, etc.), salvando em JSON."""

import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX = BASE_DIR / "2-papers" / "all_papers_llm_classif_final.xlsx"
REFS_DIR = Path(__file__).resolve().parent / "refs_por_estudo"
OUT_DIR = REFS_DIR  # JSON e TXT na mesma pasta


def load_rejected_keys() -> set[str]:
    """Retorna o conjunto de chaves (stem do PDF) rejeitadas na triagem."""
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["Classificação LLM"]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}
    triagem_idx = col["Triagem"]
    pdf_idx = col["Arquivo PDF"]
    rejected = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[triagem_idx].value == "REJEITADO":
            pdf = (row[pdf_idx].value or "")
            if pdf:
                rejected.add(Path(pdf).stem)
    wb.close()
    return rejected

# ── helpers ─────────────────────────────────────────────────────────────────

LIGATURE_MAP = {"ﬁ": "fi", "ﬂ": "fl", "ﬀ": "ff", "ﬃ": "ffi", "ﬄ": "ffl"}


def _fix(text: str) -> str:
    for lig, repl in LIGATURE_MAP.items():
        text = text.replace(lig, repl)
    return text


def _clean_line(line: str) -> str:
    """Remove page numbers and noise from extracted PDF lines."""
    s = line.strip()
    # Remove standalone page numbers
    if re.match(r"^\d{1,3}\s*$", s):
        return ""
    return s


def _join_lines(raw_text: str) -> str:
    """Join wrapped lines into continuous text, preserving paragraph breaks."""
    lines = raw_text.split("\n")
    cleaned = [_clean_line(l) for l in lines]
    # Join non-empty consecutive lines; keep blank lines as separators
    result = []
    buf = []
    for line in cleaned:
        if line == "":
            if buf:
                result.append(" ".join(buf))
                buf = []
            result.append("")
        else:
            buf.append(line)
    if buf:
        result.append(" ".join(buf))
    return "\n".join(result)


# ── splitting references ────────────────────────────────────────────────────

# ABNT-style: reference starts with UPPERCASE SURNAME (at least 2 uppercase words)
RE_ABNT_START = re.compile(
    r"^([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ][A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ\-']{1,})"
    r"(?:\s*,|\s+[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ])"
)

# APA-style: Surname, I. or Surname, I. I.
RE_APA_START = re.compile(
    r"^([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑa-záàâãéèêíïóôõöúçñ][a-záàâãéèêíïóôõöúçñ\-']+)"
    r",\s*[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ]\."
)

# Numbered: [1], [2], etc. or 1., 2., etc.
RE_NUM_START = re.compile(r"^\[?\d+[\]\.]\s+")


def _detect_style(text: str) -> str:
    """Detect the predominant citation style in the text."""
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    abnt = sum(1 for l in lines if RE_ABNT_START.match(l))
    apa = sum(1 for l in lines if RE_APA_START.match(l))
    numbered = sum(1 for l in lines if RE_NUM_START.match(l))
    counts = {"abnt": abnt, "apa": apa, "numbered": numbered}
    return max(counts, key=counts.get)


def _is_ref_start(line: str, style: str) -> bool:
    """Check if a line looks like the beginning of a new reference."""
    if not line.strip():
        return False
    s = line.strip()
    # Standalone page numbers are not starts
    if re.match(r"^\d{1,3}$", s):
        return False
    if style == "numbered":
        return bool(RE_NUM_START.match(s))
    elif style == "abnt":
        return bool(RE_ABNT_START.match(s))
    else:  # apa
        return bool(RE_APA_START.match(s))


def _split_references(text: str, style: str) -> list[str]:
    """Split continuous text into individual reference strings."""
    lines = text.split("\n")
    refs = []
    current = []

    for line in lines:
        stripped = line.strip()
        # Skip blank lines and standalone page numbers
        if not stripped or re.match(r"^\d{1,3}$", stripped):
            continue
        # Skip lines that are just the study header repeated
        if re.match(r"^[A-Za-z\s,\.]+\.$", stripped) and len(stripped) < 60:
            # Could be author name repeated as page header - skip if very short
            pass

        if _is_ref_start(stripped, style) and current:
            refs.append(" ".join(current))
            current = [stripped]
        else:
            current.append(stripped)

    if current:
        refs.append(" ".join(current))
    return refs


# ── parsing fields ──────────────────────────────────────────────────────────

def _extract_year(text: str) -> str | None:
    """Extract the most likely publication year (4 digits between 1900-2029)."""
    # Look for year in parentheses first (APA style)
    m = re.search(r"\((\d{4})\)", text)
    if m and 1900 <= int(m.group(1)) <= 2029:
        return m.group(1)
    # Look for year after comma or period
    years = re.findall(r"(?:,\s*|\.?\s+)(\d{4})(?:\s*[,.\)]|$)", text)
    for y in years:
        if 1900 <= int(y) <= 2029:
            return y
    # Fallback: any 4-digit year
    all_years = re.findall(r"\b(\d{4})\b", text)
    for y in all_years:
        if 1900 <= int(y) <= 2029:
            return y
    return None


def _extract_doi(text: str) -> str | None:
    m = re.search(r"(10\.\d{4,}/[^\s,]+)", text)
    return m.group(1).rstrip(".") if m else None


def _extract_url(text: str) -> str | None:
    m = re.search(r"(https?://[^\s>,]+)", text)
    return m.group(1).rstrip(".") if m else None


def _extract_volume_issue_pages(text: str) -> dict:
    """Extract v., n., p. (ABNT) or vol(issue), pages (APA)."""
    result = {}
    # ABNT: v. 72, n. 1, p. 1-19
    m = re.search(r"v\.\s*(\d+)", text)
    if m:
        result["volume"] = m.group(1)
    m = re.search(r"n\.\s*(\d+)", text)
    if m:
        result["issue"] = m.group(1)
    m = re.search(r"p\.\s*([\d]+[\-–]\d+|\d+)", text)
    if m:
        result["pages"] = m.group(1).replace("–", "-")

    # APA: 35(4), 559-588 or 37, 121-136
    if not result:
        m = re.search(r",\s*(\d+)\s*\((\d+)\)\s*,\s*(\d+[\-–]\d+)", text)
        if m:
            result["volume"] = m.group(1)
            result["issue"] = m.group(2)
            result["pages"] = m.group(3).replace("–", "-")
        else:
            m = re.search(r",\s*(\d+)\s*,\s*(\d+[\-–]\d+)", text)
            if m:
                result["volume"] = m.group(1)
                result["pages"] = m.group(2).replace("–", "-")
    return result


def _parse_abnt(ref_text: str) -> dict:
    """Parse an ABNT-style reference."""
    result = {"raw": ref_text}

    year = _extract_year(ref_text)
    if year:
        result["ano"] = year

    doi = _extract_doi(ref_text)
    if doi:
        result["doi"] = doi

    url = _extract_url(ref_text)
    if url:
        result["url"] = url

    vip = _extract_volume_issue_pages(ref_text)
    result.update(vip)

    # ABNT: SOBRENOME, Nome; SOBRENOME2, N. N. Título. Periódico, v.X, n.Y, p.Z, ANO.
    text = ref_text

    # Strategy: find the boundary between authors and title.
    # Authors are separated by ";". After the last author there is ". " + title.
    # Initials like "M. D. R." also contain periods, so we can't just split on ". ".
    #
    # Approach: scan for ". " boundaries. At each one, check if what follows
    # looks like a TITLE (starts with uppercase + lowercase word, length > 2 chars)
    # rather than another author initial or name.

    # A title-start looks like: capital letter followed by a lowercase letter or
    # a long word — NOT a single initial like "R." or a SURNAME pattern "SILVA"
    def _is_title_start(s: str) -> bool:
        s = s.strip()
        if not s:
            return False
        # Single initial: "R.", "M." — not a title
        if re.match(r"^[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ]\.", s):
            return False
        # All-caps word (another ABNT surname): "SILVA", "BRASIL" — not a title
        if re.match(r"^[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ]{2,}\b", s):
            return False
        # Semicolon continuation: "; SOBRENOME" — not a title
        if s.startswith(";"):
            return False
        # Portuguese/Spanish/English articles starting a title:
        # "O papel...", "A análise...", "Os efeitos...", "As políticas..."
        # "The impact...", "An analysis..."
        if re.match(
            r"^(?:O|A|Os|As|Um|Uma|The|An?|Do|Da|Dos|Das|No|Na|Nos|Nas)\s+[a-záàâãéèêíïóôõöúçñ]",
            s,
        ):
            return True
        # Looks like a title: starts with uppercase + has some lowercase text
        if re.match(r"^[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÇÑ\"\'][a-záàâãéèêíïóôõöúçñ]", s):
            return True
        # Starts with a number (e.g. law titles "Lei nº...")
        if re.match(r"^\d", s):
            return True
        # Starts with lowercase (e.g. "el papel...")
        if re.match(r"^[a-z]", s):
            return True
        return False

    # Find all ". " positions and test each one
    author_part = None
    remainder = ""
    pos = 0
    while True:
        idx = text.find(". ", pos)
        if idx == -1:
            break
        candidate_after = text[idx + 2:]
        if _is_title_start(candidate_after):
            author_part = text[:idx].strip()
            remainder = candidate_after.strip()
            break
        pos = idx + 2

    if author_part:
        result["autor"] = author_part.rstrip(".")
    else:
        # Fallback: first sentence
        parts = text.split(". ", 1)
        result["autor"] = parts[0].strip()
        remainder = parts[1].strip() if len(parts) > 1 else ""

    if remainder:
        # Title: text up to the next period followed by a journal/publisher indicator
        # Try to find title ending before a known journal pattern
        # Simple heuristic: title ends at the first period followed by a capitalized word
        # that looks like a journal name, or before "v.", "In:", publisher
        title_match = re.match(
            r"^(.+?)\.\s+"
            r"((?:[A-Z][a-záàâãéèêíïóôõöúçñ]|In:|Editora|Working|Technical|"
            r"Dispon[ií]vel|Bras[ií]lia|Rio de Janeiro|Fortaleza|S[aã]o Paulo|"
            r"Princeton|Oxford|Cambridge|London|New York|Geneva|Washington).*)",
            remainder,
            re.DOTALL,
        )
        if title_match:
            result["titulo"] = title_match.group(1).strip()
            after_title = title_match.group(2).strip()

            # Try to extract journal name (before v., n., p. or year)
            journal_match = re.match(
                r"^(.+?)(?:,\s*v\.\s*\d|,\s*\d{4}|\.\s*\d{4}|,\s*p\.)",
                after_title,
            )
            if journal_match:
                candidate = journal_match.group(1).strip().rstrip(",.")
                # Only set as journal if it looks like one (not a city or publisher)
                if len(candidate) > 3 and not re.match(
                    r"^(?:Bras[ií]lia|Rio de Janeiro|São Paulo|Fortaleza|"
                    r"Recife|Princeton|Oxford|Cambridge|London)",
                    candidate,
                ):
                    result["periodico"] = candidate
        else:
            # Title is the whole remainder (no journal found)
            result["titulo"] = remainder.rstrip(".")

    return result


def _parse_apa(ref_text: str) -> dict:
    """Parse an APA-style reference."""
    result = {"raw": ref_text}

    year = _extract_year(ref_text)
    if year:
        result["ano"] = year

    doi = _extract_doi(ref_text)
    if doi:
        result["doi"] = doi

    url = _extract_url(ref_text)
    if url:
        result["url"] = url

    vip = _extract_volume_issue_pages(ref_text)
    result.update(vip)

    # APA: Author, I., & Author, I. (Year). Title. Journal, vol(issue), pages.
    # Split at (year).
    m = re.match(r"^(.+?)\s*\(\d{4}\)\.\s*(.+)", ref_text, re.DOTALL)
    if m:
        result["autor"] = m.group(1).strip().rstrip(",. ")
        remainder = m.group(2).strip()

        # Title ends at the first period followed by a space and capital letter (journal)
        title_match = re.match(r"^(.+?)\.\s+([A-Z].+)", remainder, re.DOTALL)
        if title_match:
            result["titulo"] = title_match.group(1).strip()
            after_title = title_match.group(2).strip()

            # Journal: text before volume/pages pattern
            journal_match = re.match(
                r"^(.+?)(?:,\s*\d+\s*[\(,]|\.\s*$)", after_title
            )
            if journal_match:
                result["periodico"] = journal_match.group(1).strip().rstrip(",.")
        else:
            result["titulo"] = remainder.rstrip(".")
    else:
        # Fallback
        parts = ref_text.split(". ", 1)
        result["autor"] = parts[0].strip()
        if len(parts) > 1:
            result["titulo"] = parts[1].strip().rstrip(".")

    return result


def _parse_numbered(ref_text: str) -> dict:
    """Parse a numbered reference by stripping the number and delegating."""
    cleaned = re.sub(r"^\[?\d+[\]\.]\s*", "", ref_text)
    # Detect if the cleaned text looks ABNT or APA
    if RE_ABNT_START.match(cleaned):
        return _parse_abnt(cleaned)
    return _parse_apa(cleaned)


def parse_reference(ref_text: str, style: str) -> dict:
    """Parse a single reference string into structured fields."""
    ref_text = _fix(ref_text.strip())

    # Remove standalone page numbers embedded in text
    ref_text = re.sub(r"\s+\d{1,3}\s+", " ", ref_text)

    if style == "numbered":
        return _parse_numbered(ref_text)
    elif style == "abnt":
        return _parse_abnt(ref_text)
    else:
        return _parse_apa(ref_text)


# ── main ────────────────────────────────────────────────────────────────────

def process_file(txt_path: Path) -> dict:
    """Process a single _refs.txt file into structured JSON."""
    text = txt_path.read_text(encoding="utf-8")

    # Extract header info
    header_lines = []
    body_lines = []
    past_header = False
    for line in text.split("\n"):
        if line.startswith("====="):
            past_header = True
            continue
        if not past_header:
            header_lines.append(line)
        else:
            body_lines.append(line)

    # Parse header
    meta = {}
    for hl in header_lines:
        if hl.startswith("# Estudo #"):
            meta["estudo_num"] = hl.replace("# Estudo #", "").strip()
        elif hl.startswith("# Fonte:"):
            meta["pdf"] = hl.replace("# Fonte:", "").strip()
        elif hl.startswith("# ") and "autores" not in meta:
            content = hl[2:].strip()
            if "(" in content and ")" in content:
                meta["autores_estudo"] = content
            elif not meta.get("titulo_estudo"):
                meta["titulo_estudo"] = content

    raw_body = "\n".join(body_lines)

    # Detect style on raw lines (before joining)
    style = _detect_style(raw_body)
    # Split references line-by-line (don't pre-join, keeps ref boundaries)
    ref_strings = _split_references(raw_body, style)

    # Filter noise (too short, page numbers, repeated headers)
    ref_strings = [
        r for r in ref_strings
        if len(r) > 20
        and not re.match(r"^\d{1,3}$", r.strip())
        and not re.match(r"(?i)^refer[eê]n", r.strip())
    ]

    # Parse each reference
    parsed_refs = []
    for ref_str in ref_strings:
        parsed = parse_reference(ref_str, style)
        parsed_refs.append(parsed)

    return {
        "meta": meta,
        "estilo_detectado": style,
        "total_referencias": len(parsed_refs),
        "referencias": parsed_refs,
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    rejected = load_rejected_keys()
    all_txt = sorted(REFS_DIR.glob("*_refs.txt"))
    txt_files = [f for f in all_txt if f.stem.replace("_refs", "") not in rejected]
    skipped = len(all_txt) - len(txt_files)
    if skipped:
        print(f"Ignorando {skipped} arquivo(s) de estudos rejeitados na triagem")
    print(f"Arquivos a processar: {len(txt_files)}")

    all_results = {}
    total_refs = 0
    total_with_title = 0
    total_with_journal = 0

    for txt_path in txt_files:
        result = process_file(txt_path)
        stem = txt_path.stem.replace("_refs", "")

        # Save individual JSON
        json_path = OUT_DIR / f"{stem}_refs.json"
        json_path.write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        all_results[stem] = result
        n = result["total_referencias"]
        total_refs += n
        total_with_title += sum(
            1 for r in result["referencias"] if r.get("titulo")
        )
        total_with_journal += sum(
            1 for r in result["referencias"] if r.get("periodico")
        )

        print(f"  {stem}: {n} refs ({result['estilo_detectado']})")

    # Save consolidated JSON
    consolidated_path = Path(__file__).resolve().parent / "referencias_estruturadas.json"
    consolidated_path.write_text(
        json.dumps(all_results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"\nResumo:")
    print(f"  Estudos processados: {len(txt_files)}")
    print(f"  Total de referências: {total_refs}")
    print(f"  Com título extraído: {total_with_title} ({100*total_with_title/max(total_refs,1):.0f}%)")
    print(f"  Com periódico extraído: {total_with_journal} ({100*total_with_journal/max(total_refs,1):.0f}%)")
    print(f"\nJSON individuais: {OUT_DIR}")
    print(f"JSON consolidado: {consolidated_path}")


if __name__ == "__main__":
    main()
