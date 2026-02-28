"""Extrai a seção de referências bibliográficas dos PDFs dos estudos
aprovados na triagem e salva em arquivos individuais .txt e um consolidado."""

import re
import sys
from pathlib import Path

import openpyxl

try:
    import pymupdf as fitz  # PyMuPDF >= 1.24
except ImportError:
    import fitz  # PyMuPDF < 1.24

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX = BASE_DIR / "2-papers" / "all_papers_llm_classif_final.xlsx"
PDF_DIR = BASE_DIR / "2-papers" / "2-2-papers-pdfs"
OUT_DIR = Path(__file__).resolve().parent / "refs_por_estudo"

# Padrões para detectar início da seção de referências
REF_PATTERNS = [
    r"(?i)^\s*refer[eê]ncias?\s*bibliogr[aá]ficas?\s*$",
    r"(?i)^\s*refer[eê]ncias?\s*$",
    r"(?i)^\s*references?\s*$",
    r"(?i)^\s*bibliography\s*$",
    r"(?i)^\s*bibliogra(f|ﬁ)ia\s*$",
    r"(?i)^\s*literatura\s+citada\s*$",
    r"(?i)^\s*refer[eê]ncias?\s*bibliogr[aá]ficas?\s*\n",
    r"(?i)^\s*\d+[\.\)]\s*refer[eê]ncias?\s*$",
    r"(?i)^\s*\d+[\.\)]\s*references?\s*$",
    # LaTeX decomposed accents (e.g. Referˆencias)
    r"(?i)^\s*refer[\^ˆ]encias?\s*$",
    r"(?i)^\s*refer[\^ˆ]encias?\s*bibliogr[\´´]aficas?\s*$",
]

# Ligatures comuns em PDFs que atrapalham matching
LIGATURE_MAP = {"ﬁ": "fi", "ﬂ": "fl", "ﬀ": "ff", "ﬃ": "ffi", "ﬄ": "ffl"}

# Padrões para detectar seções que vêm DEPOIS das referências (para cortar)
POST_REF_PATTERNS = [
    r"(?i)^\s*ap[eê]ndice",
    r"(?i)^\s*anexo",
    r"(?i)^\s*appendix",
    r"(?i)^\s*notas?\s*$",
    r"(?i)^\s*agradecimentos?\s*$",
    r"(?i)^\s*acknowledgements?\s*$",
]


def _fix_ligatures(text: str) -> str:
    """Replace common PDF ligatures with their ASCII equivalents."""
    for lig, repl in LIGATURE_MAP.items():
        text = text.replace(lig, repl)
    return text


def find_ref_section(text: str) -> str | None:
    """Encontra e retorna a seção de referências do texto completo do PDF."""
    lines = text.split("\n")
    ref_start = None

    # Procurar de trás para frente (a seção de referências geralmente está no final)
    for i in range(len(lines) - 1, -1, -1):
        line = _fix_ligatures(lines[i].strip())
        if not line:
            continue
        for pat in REF_PATTERNS:
            if re.match(pat, line):
                ref_start = i
                break
        if ref_start is not None:
            break

    if ref_start is None:
        return None

    # Extrair do início da seção até o final ou até uma seção pós-referências
    ref_lines = []
    for i in range(ref_start + 1, len(lines)):
        line = lines[i]
        # Verificar se chegamos em uma seção pós-referências
        stripped = line.strip()
        is_post = False
        for pat in POST_REF_PATTERNS:
            if re.match(pat, stripped):
                is_post = True
                break
        if is_post:
            break
        ref_lines.append(line)

    # Limpar linhas em branco no início e fim
    text_refs = "\n".join(ref_lines).strip()
    return text_refs if text_refs else None


def extract_text_from_pdf(pdf_path: Path) -> str:
    """Extrai texto completo de um PDF usando PyMuPDF."""
    doc = fitz.open(str(pdf_path))
    text = ""
    for page in doc:
        text += page.get_text() + "\n"
    doc.close()
    return text


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["Classificação LLM"]
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}

    triagem_idx = col["Triagem"]
    pdf_idx = col["Arquivo PDF"]
    titulo_idx = col["Titulo"]
    autores_idx = col["Autores"]
    ano_idx = col["Ano"]
    num_idx = col["#"]

    approved = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[triagem_idx].value != "APROVADO":
            continue
        pdf_name = row[pdf_idx].value
        if not pdf_name:
            continue
        approved.append({
            "num": row[num_idx].value,
            "pdf": str(pdf_name).strip(),
            "titulo": str(row[titulo_idx].value or "").strip(),
            "autores": str(row[autores_idx].value or "").strip(),
            "ano": row[ano_idx].value,
        })
    wb.close()

    print(f"Estudos aprovados com PDF: {len(approved)}")

    success = 0
    failed = []
    all_refs = []

    for study in sorted(approved, key=lambda x: x["num"]):
        pdf_path = PDF_DIR / study["pdf"]
        if not pdf_path.exists():
            print(f"  [ERRO] PDF não encontrado: {study['pdf']}")
            failed.append((study["num"], study["pdf"], "PDF não encontrado"))
            continue

        try:
            text = extract_text_from_pdf(pdf_path)
        except Exception as e:
            print(f"  [ERRO] Falha ao ler {study['pdf']}: {e}")
            failed.append((study["num"], study["pdf"], f"Erro leitura: {e}"))
            continue

        refs = find_ref_section(text)
        if not refs:
            print(f"  [AVISO] Referências não encontradas: {study['pdf']}")
            failed.append((study["num"], study["pdf"], "Seção de referências não encontrada"))
            continue

        # Salvar arquivo individual
        out_name = Path(study["pdf"]).stem + "_refs.txt"
        out_path = OUT_DIR / out_name
        header = (
            f"# Estudo #{study['num']}\n"
            f"# {study['autores']} ({study['ano']})\n"
            f"# {study['titulo']}\n"
            f"# Fonte: {study['pdf']}\n"
            f"{'=' * 70}\n\n"
        )
        out_path.write_text(header + refs, encoding="utf-8")

        all_refs.append({
            "num": study["num"],
            "autores": study["autores"],
            "ano": study["ano"],
            "titulo": study["titulo"],
            "pdf": study["pdf"],
            "refs": refs,
        })
        success += 1

    # Salvar consolidado
    consolidated = Path(__file__).resolve().parent / "referencias_consolidadas.txt"
    with open(consolidated, "w", encoding="utf-8") as f:
        f.write(f"REFERÊNCIAS BIBLIOGRÁFICAS DOS ESTUDOS APROVADOS\n")
        f.write(f"Total de estudos processados: {success}/{len(approved)}\n")
        f.write(f"{'=' * 70}\n\n")
        for entry in all_refs:
            f.write(f"{'#' * 70}\n")
            f.write(f"# Estudo #{entry['num']} - {entry['autores']} ({entry['ano']})\n")
            f.write(f"# {entry['titulo']}\n")
            f.write(f"# PDF: {entry['pdf']}\n")
            f.write(f"{'#' * 70}\n\n")
            f.write(entry["refs"])
            f.write(f"\n\n")

    # Relatório
    print(f"\nResultado:")
    print(f"  Extraídos com sucesso: {success}")
    print(f"  Falhas: {len(failed)}")
    if failed:
        print(f"\nDetalhes das falhas:")
        for num, pdf, reason in failed:
            print(f"  #{num} {pdf}: {reason}")
    print(f"\nArquivos salvos em: {OUT_DIR}")
    print(f"Consolidado: {consolidated}")


if __name__ == "__main__":
    main()
