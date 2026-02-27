#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
status_pdfs.py - Resumo do status de download dos PDFs em all_papers.xlsx.

Le a planilha "Registros" de all_papers.xlsx e imprime:
  1. Tabela-resumo por Base (Total | Baixado | Faltando | % Baixado)
  2. Listagem detalhada dos registros com Baixado="Nao", agrupados por Base.

Uso:  python -X utf8 status_pdfs.py
"""

import io
import shutil
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuracao
# ---------------------------------------------------------------------------
XLSX = Path(__file__).with_name("all_papers.xlsx")
SHEET = "Registros"

# Colunas (0-indexed apos leitura como tupla)
COL_BASE     = 0   # A
COL_URL      = 1   # B
COL_BAIXADO  = 2   # C
COL_ARQUIVO  = 3   # D
COL_ID       = 4   # E
COL_TITULO   = 5   # F
COL_AUTORES  = 6   # G
COL_ANO      = 7   # H
COL_PERIODICO= 8   # I
COL_DOI      = 9   # J
COL_RESUMO   = 10  # K
COL_TIPO     = 11  # L
COL_PALCHAVE = 12  # M
COL_OBS      = 13  # N


def trunc(text, maxlen):
    """Trunca string; retorna '' se None."""
    if text is None:
        return ""
    s = str(text)
    return s if len(s) <= maxlen else s[: maxlen - 3] + "..."


def open_workbook(path):
    """Tenta abrir o workbook; se falhar por permissao, le via bytes em memoria."""
    try:
        return openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except PermissionError:
        pass

    # Fallback: ler os bytes brutos (funciona mesmo com arquivo aberto no Excel)
    try:
        with open(str(path), "rb") as f:
            data = f.read()
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except PermissionError:
        pass

    # Ultimo recurso: tentar uma copia pre-existente
    copy_path = path.with_name(path.stem + "_copy" + path.suffix)
    if copy_path.exists():
        print(f"[AVISO] Usando copia existente: {copy_path.name}")
        return openpyxl.load_workbook(str(copy_path), read_only=True, data_only=True)

    # Tenta criar uma copia via shutil
    try:
        tmp = path.with_name("_tmp_status_copy.xlsx")
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(str(tmp), read_only=True, data_only=True)
        return wb
    except PermissionError:
        print(f"ERRO: Nao foi possivel abrir {path.name}.")
        print("       Feche o arquivo no Excel e tente novamente.")
        sys.exit(1)


def main():
    wb = open_workbook(XLSX)
    ws = wb[SHEET]

    # ------------------------------------------------------------------
    # Leitura dos dados
    # ------------------------------------------------------------------
    records = []  # lista de tuplas (row_number, valores)
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        base = row[COL_BASE]
        if base is None:
            continue  # pula linhas vazias
        records.append((idx, row))

    wb.close()

    # Limpar copia temporaria se existir
    tmp = XLSX.with_name("_tmp_status_copy.xlsx")
    if tmp.exists():
        try:
            tmp.unlink()
        except OSError:
            pass

    # ------------------------------------------------------------------
    # Agrupamento por Base
    # ------------------------------------------------------------------
    bases = OrderedDict()  # base -> {"total", "sim", "nao", "missing"}
    for row_num, vals in records:
        base = str(vals[COL_BASE]).strip()
        baixado = str(vals[COL_BAIXADO] or "").strip()

        if base not in bases:
            bases[base] = {"total": 0, "sim": 0, "nao": 0, "missing": []}

        bases[base]["total"] += 1
        if baixado == "Sim":
            bases[base]["sim"] += 1
        else:
            bases[base]["nao"] += 1
            bases[base]["missing"].append((row_num, vals))

    # ------------------------------------------------------------------
    # Impressao da tabela-resumo
    # ------------------------------------------------------------------
    sep = "-" * 72
    header_fmt = "{:<16s} {:>7s} {:>9s} {:>9s} {:>10s}"
    row_fmt    = "{:<16s} {:>7d} {:>9d} {:>9d} {:>9.1f}%"

    print()
    print("=" * 72)
    print("  STATUS DOS PDFs  -  all_papers.xlsx / Registros")
    print("=" * 72)
    print()
    print(header_fmt.format("Base", "Total", "Baixado", "Faltando", "% Baixado"))
    print(sep)

    tot_total = tot_sim = tot_nao = 0
    for base, d in bases.items():
        pct = 100 * d["sim"] / d["total"] if d["total"] else 0
        print(row_fmt.format(base, d["total"], d["sim"], d["nao"], pct))
        tot_total += d["total"]
        tot_sim   += d["sim"]
        tot_nao   += d["nao"]

    print(sep)
    pct_total = 100 * tot_sim / tot_total if tot_total else 0
    print(row_fmt.format("TOTAL", tot_total, tot_sim, tot_nao, pct_total))
    print()

    # ------------------------------------------------------------------
    # Listagem dos registros faltantes
    # ------------------------------------------------------------------
    any_missing = any(d["nao"] > 0 for d in bases.values())
    if not any_missing:
        print("Todos os PDFs foram baixados!")
        return

    print("=" * 72)
    print("  REGISTROS COM PDF FALTANDO")
    print("=" * 72)

    detail_hdr = "  {:<5s} {:<5s} {:<50s} {:<70s} {:<s}"
    detail_row = "  {:<5s} {:<5s} {:<50s} {:<70s} {:<s}"

    for base, d in bases.items():
        if d["nao"] == 0:
            continue

        print()
        print(f"  >>> {base}  ({d['nao']} faltando)")
        print(detail_hdr.format("Row", "Ano", "Autores", "Titulo", "DOI"))
        print("  " + "-" * 140)

        for row_num, vals in d["missing"]:
            ano     = str(vals[COL_ANO] or "")
            autores = trunc(vals[COL_AUTORES], 50)
            titulo  = trunc(vals[COL_TITULO], 70)
            doi     = str(vals[COL_DOI] or "")
            print(detail_row.format(str(row_num), ano, autores, titulo, doi))

    print()


if __name__ == "__main__":
    main()
