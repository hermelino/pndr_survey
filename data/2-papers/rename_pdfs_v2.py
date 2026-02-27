"""
Re-rename all PDFs in 2-papers root with improved author extraction.
Convention: <base>-<year>-<surname1>-<surname2>-<surname3>.pdf

Improvements:
- Detects CAPES "FullName, FullName" format vs Scopus "Surname, First" format
- Extracts only the LAST word of each surname (e.g. "Mendes Resende" -> "resende")
- Much better ANPEC PDF author extraction
- Skips prepositions (de, da, do, dos, das) at end of names
"""
import os, re, sys, shutil, unicodedata
import openpyxl
import pdfplumber

ROOT = 'C:/OneDrive/github/pndr_survey/data/2-papers'
XLSX = os.path.join(ROOT, 'all_records.xlsx')
SRC_XLSX = 'C:/OneDrive/github/pndr_survey/data/1-records/all_records.xlsx'

PREPOSITIONS = {'de', 'da', 'do', 'dos', 'das', 'del', 'di', 'e', 'a', 'o', 'os', 'as'}

# ── Utilities ──────────────────────────────────────────────────────────

def remove_accents(s):
    s = unicodedata.normalize('NFD', s)
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def clean_surname(s):
    """Extract the last meaningful word from a surname, lowercased, no accents."""
    s = remove_accents(s).lower().strip()
    # Remove any non-alpha chars
    tokens = re.findall(r'[a-z]+', s)
    if not tokens:
        return ''
    # Walk backwards to find last non-preposition word
    for t in reversed(tokens):
        if t not in PREPOSITIONS and len(t) >= 2:
            return t
    # Fallback: return last token
    return tokens[-1] if tokens else ''


def detect_author_format(authors_str):
    """Detect if authors use 'Surname, First; Surname2, First2' (semicolon-based)
    or 'First Last, First2 Last2, ...' (comma-based full names)."""
    if not authors_str:
        return 'empty'
    if ';' in authors_str:
        return 'semicolon'  # Scopus/EconPapers: "Surname, First; Surname2, First2"
    # Check if commas separate full names or "Surname, First" pairs
    # CAPES typically has full names like "João Silva, Maria Santos"
    parts = [p.strip() for p in authors_str.split(',')]
    if len(parts) >= 2:
        # If most parts have 2+ words each, it's full names
        multi_word = sum(1 for p in parts if len(p.split()) >= 2)
        if multi_word >= len(parts) * 0.5:
            return 'comma_fullnames'  # "Full Name, Full Name, Full Name"
        else:
            return 'semicolon'  # Single entry "Surname, FirstName"
    return 'single'


def parse_surnames(authors_str):
    """Parse author surnames from any format. Returns list of clean surnames."""
    if not authors_str or not authors_str.strip():
        return []

    fmt = detect_author_format(authors_str)
    surnames = []

    if fmt == 'semicolon':
        # "Surname, First; Surname2, First2" or "First Last; First2 Last2"
        for part in re.split(r';\s*', authors_str):
            part = part.strip()
            if not part:
                continue
            if ',' in part:
                # "Surname, First" -> take surname part
                surname_part = part.split(',')[0].strip()
            else:
                # "First Last" -> take last word
                surname_part = part
            s = clean_surname(surname_part)
            if s and len(s) >= 2:
                surnames.append(s)

    elif fmt == 'comma_fullnames':
        # "Full Name1, Full Name2, Full Name3" -> take last word of each
        parts = [p.strip() for p in authors_str.split(',')]
        for part in parts:
            if not part:
                continue
            s = clean_surname(part)
            if s and len(s) >= 2:
                surnames.append(s)

    elif fmt == 'single':
        s = clean_surname(authors_str)
        if s and len(s) >= 2:
            surnames.append(s)

    return surnames


# ── PDF author extraction ──────────────────────────────────────────────

INSTITUTION_WORDS = {
    'universidade', 'federal', 'estadual', 'instituto', 'faculdade', 'programa',
    'departamento', 'professor', 'doutor', 'mestr', 'doutorando', 'pesquisador',
    'cedeplar', 'ipea', 'ufrj', 'ufmg', 'ufpe', 'ufc', 'caen', 'ufba', 'uff',
    'unifacs', 'unifa', 'ufrgs', 'unicamp', 'ufal', 'ufopa', 'ppger',
    'curso', 'phd', 'adjunto', 'associado', 'titular', 'pos', 'graduacao',
    'mestrado', 'doutorado', 'email', 'telefone', 'endereco', 'filiacao',
    'area', 'sessao', 'mesa', 'palavras', 'keyword', 'jel', 'resumo', 'abstract',
    'economia', 'regional', 'urbana', 'applied', 'economics', 'encontro',
    'nacional', 'prêmio', 'premio', 'programacao', 'caixa', 'banco',
    'salao', 'sala', 'hotel', 'resort', 'abertura', 'inscricoes', 'inscricao',
    'sessão', 'presidente', 'secretaria', 'tesouraria', 'organizacao',
    'patrocinadores', 'rua', 'avenida', 'cep', 'benfica', 'fortaleza',
    'salvador', 'natal', 'rio', 'janeiro', 'paulo', 'brasilia',
    'selecionados', 'artigos', 'resultado', 'lugar', 'vencedor',
}


def is_likely_name(text):
    """Check if text looks like a person name (not institution/place)."""
    text_lower = remove_accents(text).lower()
    words = text_lower.split()
    if not words:
        return False
    # Check if any word is an institution word
    for w in words:
        w_clean = re.sub(r'[^a-z]', '', w)
        if w_clean in INSTITUTION_WORDS:
            return False
    # Names are typically 2-5 words, start with uppercase
    if not (2 <= len(words) <= 5):
        return False
    return True


def extract_authors_from_pdf(pdf_path):
    """Extract author surnames from first page of PDF."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return []
            txt = pdf.pages[0].extract_text() or ''
    except Exception:
        return []

    lines = [l.strip() for l in txt.split('\n') if l.strip()]
    if len(lines) < 2:
        return []

    # Strategy 1: Dash-separated author line
    for i, line in enumerate(lines[:15]):
        if re.search(r'\s+[-–]\s+', line):
            parts = re.split(r'\s*[-–]\s*', line)
            names = [p.strip() for p in parts if p.strip()]
            # Check all parts look like names
            if len(names) >= 2 and all(is_likely_name(n) for n in names):
                surnames = [clean_surname(n) for n in names]
                surnames = [s for s in surnames if s and len(s) >= 2]
                if len(surnames) >= 2:
                    return surnames

    # Strategy 2: "Autor(es):" label
    for i, line in enumerate(lines[:25]):
        m = re.match(r'^\s*auto(r|res)\s*:\s*(.*)', line, re.I)
        if m:
            author_text = m.group(2).strip()
            if not author_text and i + 1 < len(lines):
                author_text = lines[i + 1].strip()
            if author_text:
                # Remove affiliation after dash
                author_text = re.split(r'\s*[-–]\s*(?:Doutor|Mestr|Professor|Programa|Universidade)', author_text, flags=re.I)[0]
                # Try splitting by comma, "e", "and"
                parts = re.split(r',\s*|\s+e\s+|\s+and\s+', author_text)
                surnames = []
                for p in parts:
                    p = p.strip()
                    if p and is_likely_name(p):
                        s = clean_surname(p)
                        if s and len(s) >= 2:
                            surnames.append(s)
                if surnames:
                    return surnames

    # Strategy 3: "Titulo:"/"Autor:" explicit format
    for i, line in enumerate(lines[:15]):
        if re.match(r'^\s*t[ií]tulo\s*:', line, re.I):
            for j in range(i + 1, min(i + 5, len(lines))):
                m2 = re.match(r'^\s*auto(r|res?)\s*:\s*(.*)', lines[j], re.I)
                if m2:
                    name_text = m2.group(2).strip()
                    if name_text:
                        s = clean_surname(name_text)
                        if s and len(s) >= 2:
                            return [s]

    # Strategy 4: Names between title and RESUMO/Abstract
    resumo_idx = None
    for i, line in enumerate(lines[:35]):
        if re.match(r'^\s*(resumo|abstract)\s*:?\s*$', line, re.I):
            resumo_idx = i
            break

    if resumo_idx and resumo_idx > 2:
        # Find where title ends (first few lines are usually title)
        # Title lines tend to be ALL CAPS or long
        title_end = 0
        for i in range(min(6, resumo_idx)):
            line = lines[i]
            # Skip if it looks like a title line (long, all caps, or has title markers)
            if len(line) > 60 or line.isupper() or re.match(r'^(area|\d)', line, re.I):
                title_end = i + 1
            else:
                break

        # Collect candidate name lines between title_end and resumo
        names = []
        for i in range(max(1, title_end), resumo_idx):
            line = lines[i]
            # Remove footnote markers
            clean_line = re.sub(r'[*†‡§¤"∗\d]+$', '', line).strip()
            clean_line = re.sub(r'^[\d*†‡§¤"∗]+\s*', '', clean_line).strip()
            if not clean_line:
                continue
            if is_likely_name(clean_line):
                names.append(clean_line)

        if names:
            surnames = [clean_surname(n) for n in names]
            surnames = [s for s in surnames if s and len(s) >= 2]
            if surnames:
                return surnames

    return []


def extract_year_from_pdf(pdf_path):
    """Try to extract publication year from PDF first page."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None
            txt = pdf.pages[0].extract_text() or ''
    except Exception:
        return None
    m = re.search(r'(?:19|20)\d{2}', txt[:1000])
    if m:
        yr = int(m.group())
        if 1990 <= yr <= 2026:
            return yr
    return None


def build_filename(base, year, surnames):
    """Build: base-year-surname1-surname2-surname3.pdf"""
    parts = [base]
    parts.append(str(year) if year else 'nd')
    for s in surnames[:3]:
        parts.append(s)
    name = '-'.join(parts)
    return name + '.pdf'


# ── Read ORIGINAL xlsx to get clean author data ───────────────────────

wb_src = openpyxl.load_workbook(SRC_XLSX)
ws_src = wb_src['Registros']

# Read updated xlsx to get current PDF filenames
wb = openpyxl.load_workbook(XLSX)
ws = wb['Registros']

# Build mapping: current_pdf_name -> record info (from ORIGINAL authors)
pdf_to_record = {}

for row in range(2, ws.max_row + 1):
    baixado = ws.cell(row=row, column=12).value
    if baixado != 'Sim':
        continue

    pdf = ws.cell(row=row, column=13).value or ''
    if not pdf:
        continue

    # Get original authors from source xlsx (same row structure)
    base = ws_src.cell(row=row, column=1).value or ''
    authors = ws_src.cell(row=row, column=4).value or ''
    year = ws_src.cell(row=row, column=5).value

    if pdf not in pdf_to_record:
        pdf_to_record[pdf] = {
            'row': row,
            'base': base,
            'authors': authors,
            'year': year,
        }

print(f'Records mapped to PDFs: {len(pdf_to_record)}')

# ── Process ALL current PDFs ──────────────────────────────────────────

current_pdfs = [f for f in os.listdir(ROOT) if f.lower().endswith('.pdf')]
print(f'PDFs in root: {len(current_pdfs)}')

renames = []  # (old_name, new_name)

for fname in sorted(current_pdfs):
    full_path = os.path.join(ROOT, fname)
    rec = pdf_to_record.get(fname)

    if rec:
        base = rec['base']
        year = rec['year']
        surnames = parse_surnames(rec['authors'])

        if not surnames:
            surnames = extract_authors_from_pdf(full_path)

        if not year:
            year = extract_year_from_pdf(full_path)
    else:
        # Unmapped PDF - determine base from current name prefix
        if fname.startswith('anpec-'):
            base = 'anpec'
        elif fname.startswith('econpapers-'):
            base = 'econpapers'
        elif fname.startswith('scopus-'):
            base = 'scopus'
        elif fname.startswith('scielo-'):
            base = 'scielo'
        elif fname.startswith('capes-'):
            base = 'capes'
        else:
            base = 'capes'

        surnames = extract_authors_from_pdf(full_path)
        year = extract_year_from_pdf(full_path)

    new_name = build_filename(base, year, surnames)
    renames.append((fname, new_name))

# ── Handle collisions ─────────────────────────────────────────────────

name_counts = {}
final_renames = []

for old_name, new_name in renames:
    base_name = new_name[:-4]
    if new_name in name_counts:
        name_counts[new_name] += 1
        suffix = chr(ord('a') + name_counts[new_name] - 1)
        new_name = f'{base_name}{suffix}.pdf'
    else:
        name_counts[new_name] = 1
    final_renames.append((old_name, new_name))

# ── Print plan ─────────────────────────────────────────────────────────

print('\n' + '=' * 80)
print('RENAME PLAN (v2)')
print('=' * 80)

for old_name, new_name in final_renames:
    if old_name != new_name:
        print(f'  {old_name}')
        print(f'    -> {new_name}')
    else:
        print(f'  {old_name}  (unchanged)')

changes = sum(1 for o, n in final_renames if o != n)
print(f'\nTotal: {len(final_renames)} files, {changes} to rename')

# ── Execute ────────────────────────────────────────────────────────────

print('\n' + '=' * 80)
print('EXECUTING')
print('=' * 80)

# Two-phase rename to avoid conflicts (rename to temp first, then to final)
temp_map = {}
for old_name, new_name in final_renames:
    if old_name != new_name:
        temp_name = f'_tmp_{hash(old_name) & 0xFFFF:04x}_{new_name}'
        old_path = os.path.join(ROOT, old_name)
        temp_path = os.path.join(ROOT, temp_name)
        try:
            os.rename(old_path, temp_path)
            temp_map[temp_name] = new_name
        except Exception as e:
            print(f'  ERROR (phase 1): {old_name} -> {e}')

for temp_name, new_name in temp_map.items():
    temp_path = os.path.join(ROOT, temp_name)
    new_path = os.path.join(ROOT, new_name)
    try:
        os.rename(temp_path, new_path)
        print(f'  OK: {new_name}')
    except Exception as e:
        print(f'  ERROR (phase 2): {temp_name} -> {new_name}: {e}')

# ── Update xlsx ────────────────────────────────────────────────────────

print('\n' + '=' * 80)
print('UPDATING XLSX')
print('=' * 80)

old_to_new = {o: n for o, n in final_renames if o != n}

updated = 0
for row in range(2, ws.max_row + 1):
    pdf = ws.cell(row=row, column=13).value or ''
    if pdf in old_to_new:
        ws.cell(row=row, column=13, value=old_to_new[pdf])
        updated += 1

# Also update Arquivos Extra
if 'Arquivos Extra' in wb.sheetnames:
    ws_extra = wb['Arquivos Extra']
    for row in range(2, ws_extra.max_row + 1):
        fname = ws_extra.cell(row=row, column=2).value or ''
        if fname in old_to_new:
            ws_extra.cell(row=row, column=2, value=old_to_new[fname])

wb.save(XLSX)
print(f'Updated {updated} rows')

print('\nDONE!')
