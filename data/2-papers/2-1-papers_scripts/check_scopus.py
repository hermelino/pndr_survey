"""
check_scopus.py
Lists all Scopus records from all_papers.xlsx (sheet "Registros")
and highlights those not yet downloaded (Baixado != "Sim").
"""

import openpyxl
from pathlib import Path

XLSX = Path(__file__).with_name("all_papers.xlsx")

wb = openpyxl.load_workbook(XLSX)
ws = wb["Registros"]

# Column mapping (1-based)
COL_BASE = 1        # A
COL_BAIXADO = 3     # C
COL_ARQUIVO = 4     # D
COL_TITULO = 6      # F
COL_AUTORES = 7     # G
COL_ANO = 8         # H
COL_DOI = 10        # J

scopus_rows = []
not_downloaded = []

for row_idx in range(2, ws.max_row + 1):
    base = ws.cell(row_idx, COL_BASE).value
    if base and str(base).strip().lower() == "scopus":
        titulo   = ws.cell(row_idx, COL_TITULO).value or ""
        autores  = ws.cell(row_idx, COL_AUTORES).value or ""
        ano      = ws.cell(row_idx, COL_ANO).value or ""
        doi      = ws.cell(row_idx, COL_DOI).value or ""
        baixado  = ws.cell(row_idx, COL_BAIXADO).value or ""
        arquivo  = ws.cell(row_idx, COL_ARQUIVO).value or ""

        rec = {
            "row": row_idx,
            "titulo": str(titulo).strip(),
            "autores": str(autores).strip(),
            "ano": str(ano).strip(),
            "doi": str(doi).strip(),
            "baixado": str(baixado).strip(),
            "arquivo": str(arquivo).strip(),
        }
        scopus_rows.append(rec)
        if rec["baixado"] != "Sim":
            not_downloaded.append(rec)

# -- Section 1: full list (compact) -------------------------------------------
print("=" * 100)
print(f"ALL SCOPUS RECORDS  ({len(scopus_rows)} total)")
print("=" * 100)

for i, r in enumerate(scopus_rows, 1):
    tit_short = r["titulo"][:80] + ("..." if len(r["titulo"]) > 80 else "")
    aut_short = r["autores"][:50] + ("..." if len(r["autores"]) > 50 else "")
    arq_short = r["arquivo"][:40] if r["arquivo"] else ""
    doi_short = r["doi"][:40] if r["doi"] else ""
    print(
        f"  {i:3d}. [Row {r['row']:3d}]  "
        f"Ano={r['ano']:<5s}  Baixado={r['baixado']:<4s}  "
        f"PDF={arq_short:<40s}  "
        f"DOI={doi_short}"
    )
    print(f"       Titulo : {tit_short}")
    print(f"       Autores: {aut_short}")
    print()

# -- Section 2: NOT downloaded (full details) ----------------------------------
print()
print("=" * 100)
print(f"SCOPUS RECORDS WHERE Baixado != 'Sim'  ({len(not_downloaded)} records)")
print("=" * 100)

if not not_downloaded:
    print("  (none -- all Scopus records are marked as downloaded)")
else:
    for i, r in enumerate(not_downloaded, 1):
        print(f"\n  {i}. [Row {r['row']}]  Ano={r['ano']}  Baixado=\"{r['baixado']}\"  DOI={r['doi']}")
        print(f"     Titulo  : {r['titulo']}")
        print(f"     Autores : {r['autores']}")
        print(f"     Arquivo : {r['arquivo']}")

print()
print("--- Done ---")
