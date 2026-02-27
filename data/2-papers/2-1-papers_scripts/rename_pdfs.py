"""
Rename and flatten all PDFs in 2-papers.
Convention: <base>-<year>-<surname1>-<surname2>-<surname3>.pdf
- All lowercase, no accents
- Up to 3 author surnames
- If no year: "nd"
- Handles duplicates (same physical file mapped by multiple records)
"""
import os, re, sys, shutil, unicodedata
import openpyxl
import pdfplumber

ROOT = 'C:/OneDrive/github/pndr_survey/data/2-papers'
XLSX = os.path.join(ROOT, 'all_records.xlsx')
ANPEC_DIR = os.path.join(ROOT, '2-5-papers-anpec')
ECON_DIR = os.path.join(ROOT, '2-4-papers-econpapers')

# ── Utilities ──────────────────────────────────────────────────────────

def remove_accents(s):
    """Remove accents and normalize to ASCII."""
    s = unicodedata.normalize('NFD', s)
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def clean_name(s):
    """Normalize a surname for filename use."""
    s = remove_accents(s).lower().strip()
    s = re.sub(r'[^a-z]', '', s)
    return s


def parse_surnames_from_field(authors_str):
    """Parse author surnames from xlsx Authors field.
    Formats: 'Surname1, First; Surname2, First' or 'First Last; First Last'
    """
    if not authors_str or not authors_str.strip():
        return []
    surnames = []
    for part in re.split(r';\s*', authors_str):
        part = part.strip()
        if not part:
            continue
        if ',' in part:
            # "Surname, FirstName" format
            surname = part.split(',')[0].strip()
        else:
            # "FirstName LastName" format
            tokens = part.split()
            if tokens:
                surname = tokens[-1]
            else:
                continue
        s = clean_name(surname)
        if len(s) >= 2:
            surnames.append(s)
    return surnames


def extract_authors_from_pdf(pdf_path):
    """Extract author surnames from first page of PDF using heuristics."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return []
            txt = pdf.pages[0].extract_text() or ''
    except Exception:
        return []

    lines = txt.split('\n')
    if len(lines) < 2:
        return []

    # Strategy 1: Look for dash-separated author line (common in ANPEC)
    # Pattern: "Name1 - Name2 - Name3" or "Name1 – Name2 – Name3"
    for i, line in enumerate(lines[:15]):
        # Must have at least one separator
        if re.search(r'\s+[-–]\s+', line) and not re.search(r'(resumo|abstract|palavras|keyword|jel|area|sessao|mesa|presidente)', line, re.I):
            parts = re.split(r'\s*[-–]\s*', line.strip())
            # Each part should look like a name (2-4 words, starts with uppercase)
            names = []
            for p in parts:
                p = p.strip()
                words = p.split()
                if 1 <= len(words) <= 6 and any(c.isupper() for c in p[:3] if c.isalpha()):
                    # Check it's not an institution or keyword
                    if not re.search(r'(universidade|ufmg|ufrj|ipea|cedeplar|ufpe|ufc|caen|programa|departamento|professor|doutor|mestr)', p, re.I):
                        names.append(p)
                    elif len(words) <= 3:
                        # Could still be a name even if contains abbreviation
                        pass
            if len(names) >= 2:
                surnames = []
                for n in names:
                    tokens = n.split()
                    surname = clean_name(tokens[-1])
                    if len(surname) >= 2:
                        surnames.append(surname)
                if surnames:
                    return surnames

    # Strategy 2: Look for "Autor(es):" line
    for i, line in enumerate(lines[:20]):
        if re.match(r'^\s*auto(r|res)\s*:', line, re.I):
            # Next content has author names
            author_text = line.split(':', 1)[1].strip()
            if not author_text and i + 1 < len(lines):
                author_text = lines[i + 1].strip()
            if author_text:
                # Try splitting by comma or "and"/e
                parts = re.split(r',\s*|\s+e\s+', author_text)
                surnames = []
                for p in parts:
                    p = p.strip()
                    tokens = p.split()
                    if tokens:
                        s = clean_name(tokens[-1])
                        if len(s) >= 2:
                            surnames.append(s)
                if surnames:
                    return surnames

    # Strategy 3: Look for "Titulo:"/"Autor:" labeled format
    for i, line in enumerate(lines[:15]):
        m = re.match(r'^\s*t[ií]tulo\s*:', line, re.I)
        if m:
            # Look for "Autor:" after
            for j in range(i + 1, min(i + 5, len(lines))):
                m2 = re.match(r'^\s*auto(r|res)\s*:\s*(.*)', lines[j], re.I)
                if m2:
                    author_text = m2.group(2).strip()
                    if author_text:
                        tokens = author_text.split()
                        surname = clean_name(tokens[-1])
                        if len(surname) >= 2:
                            return [surname]

    # Strategy 4: Look for names between title block and RESUMO/Abstract
    # Find the RESUMO line
    resumo_idx = None
    for i, line in enumerate(lines[:30]):
        if re.match(r'^\s*(resumo|abstract)\s*:?\s*$', line.strip(), re.I):
            resumo_idx = i
            break

    if resumo_idx and resumo_idx > 2:
        # Lines between ~line 1 and RESUMO might contain authors
        # Skip lines that look like title (all caps or very long)
        candidate_lines = []
        for i in range(1, resumo_idx):
            line = lines[i].strip()
            if not line:
                continue
            # Skip if looks like title (all caps, very long, starts with number)
            if len(line) > 80:
                continue
            if re.match(r'^\d', line):
                continue
            if re.match(r'^(area|sessao|mesa|palavras|keyword|jel|email|e-mail|telefone|endereco|filiacao)', line, re.I):
                continue
            # Skip pure institution lines
            if re.match(r'^(universidade|professor|doutor|mestr|programa|departamento|cedeplar|ipea|inst)', line, re.I):
                continue
            # Check if it looks like a name (mostly title case, 2-5 words)
            words = line.split()
            if 1 <= len(words) <= 6:
                alpha_chars = sum(1 for c in line if c.isalpha())
                if alpha_chars > len(line) * 0.5:
                    # Might be a name or affiliation
                    # Names typically start with uppercase
                    if line[0].isupper() or any(c.isupper() for c in line[:5]):
                        candidate_lines.append(line)

        # Filter candidates: names typically don't have certain patterns
        names = []
        for cl in candidate_lines:
            # Remove footnote markers (asterisks, daggers, numbers)
            cl_clean = re.sub(r'[*†‡§¤"∗\d]+', '', cl).strip()
            if not cl_clean:
                continue
            # Skip if looks like institution/affiliation
            if re.search(r'(universidade|professor|doutor|mestr|programa|departamento|cedeplar|ipea|ufrj|ufmg|ufpe|ufc|caen|pesquisador|instituto|faculdade|curso|phd|pos\s)', cl_clean, re.I):
                continue
            words = cl_clean.split()
            if 2 <= len(words) <= 5:
                names.append(cl_clean)

        if names:
            surnames = []
            for n in names:
                tokens = n.split()
                s = clean_name(tokens[-1])
                if len(s) >= 2:
                    surnames.append(s)
            if surnames:
                return surnames

    # Strategy 5: For "Autores:" in award announcements
    for i, line in enumerate(lines[:30]):
        m = re.match(r'^\s*autores?\s*:\s*(.*)', line, re.I)
        if m:
            author_text = m.group(1).strip()
            if author_text:
                # Remove affiliation parts after " - "
                author_text = re.split(r'\s*-\s*(?:Doutor|Mestr|Professor|Programa)', author_text)[0]
                tokens = author_text.split()
                if tokens:
                    s = clean_name(tokens[-1])
                    if len(s) >= 2:
                        return [s]

    return []


def extract_year_from_pdf(pdf_path):
    """Try to extract publication year from PDF first page."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return None
            txt = pdf.pages[0].extract_text() or ''
    except Exception:
        return None
    # Look for 4-digit year in common patterns
    m = re.search(r'(?:19|20)\d{2}', txt[:1000])
    if m:
        yr = int(m.group())
        if 1990 <= yr <= 2026:
            return yr
    return None


def build_filename(base, year, surnames):
    """Build the new filename: base-year-surname1-surname2-surname3.pdf"""
    parts = [base]
    parts.append(str(year) if year else 'nd')
    # Take up to 3 surnames
    for s in surnames[:3]:
        parts.append(s)
    name = '-'.join(parts)
    return name + '.pdf'


# ── Read xlsx records ──────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX)
ws = wb['Registros']

# Build mapping: (filename, subpasta) -> record info (use first match)
# Also track unique physical files
file_to_record = {}  # (subpasta, filename) -> record dict
all_mapped_files = set()  # (subpasta, filename) tuples

for row in range(2, ws.max_row + 1):
    baixado = ws.cell(row=row, column=12).value
    if baixado != 'Sim':
        continue

    base = ws.cell(row=row, column=1).value or ''
    title = ws.cell(row=row, column=3).value or ''
    authors = ws.cell(row=row, column=4).value or ''
    year = ws.cell(row=row, column=5).value
    pdf = ws.cell(row=row, column=13).value or ''
    subpasta = ws.cell(row=row, column=14).value or ''

    if not pdf:
        continue

    key = (subpasta, pdf)
    all_mapped_files.add(key)

    if key not in file_to_record:
        file_to_record[key] = {
            'row': row,
            'base': base,
            'title': title,
            'authors': authors,
            'year': year,
            'pdf': pdf,
            'subpasta': subpasta,
        }

# ── Discover ALL PDFs ──────────────────────────────────────────────────

all_pdfs = []  # (full_path, subpasta, filename)

# Subfolders
for subdir_name in ['2-4-papers-econpapers', '2-5-papers-anpec']:
    subdir_path = os.path.join(ROOT, subdir_name)
    if os.path.isdir(subdir_path):
        for fname in os.listdir(subdir_path):
            if fname.lower().endswith('.pdf'):
                all_pdfs.append((os.path.join(subdir_path, fname), subdir_name, fname))

# Root
for fname in os.listdir(ROOT):
    fpath = os.path.join(ROOT, fname)
    if os.path.isfile(fpath) and fname.lower().endswith('.pdf'):
        all_pdfs.append((fpath, '', fname))

print(f'Total PDFs discovered: {len(all_pdfs)}')
print(f'Mapped to records: {len(file_to_record)}')

# ── Build rename mapping ──────────────────────────────────────────────

renames = []  # (old_full_path, new_filename, record_or_none)

# Track files already processed to avoid duplicates
processed_paths = set()

for full_path, subpasta, fname in all_pdfs:
    real_path = os.path.realpath(full_path)
    if real_path in processed_paths:
        continue
    processed_paths.add(real_path)

    key = (subpasta, fname)
    rec = file_to_record.get(key)

    if rec:
        base = rec['base']
        year = rec['year']
        authors_str = rec['authors']
        surnames = parse_surnames_from_field(authors_str)

        # If no authors in metadata (e.g., ANPEC), extract from PDF
        if not surnames:
            surnames = extract_authors_from_pdf(full_path)

        # If still no year
        if not year:
            year = extract_year_from_pdf(full_path)

        new_name = build_filename(base, year, surnames)
        renames.append((full_path, new_name, rec))
    else:
        # Unmapped PDF - try to determine base and extract info
        if subpasta == '2-5-papers-anpec':
            base = 'anpec'
        elif subpasta == '2-4-papers-econpapers':
            base = 'econpapers'
        else:
            # Root file - try to determine base from content/name
            # These are mostly CAPES or manual downloads
            base = 'capes'  # default for root unmapped files

        surnames = extract_authors_from_pdf(full_path)
        year = extract_year_from_pdf(full_path)

        new_name = build_filename(base, year, surnames)
        renames.append((full_path, new_name, None))

# ── Handle filename collisions ─────────────────────────────────────────

name_counts = {}
final_renames = []

for old_path, new_name, rec in renames:
    base_name = new_name[:-4]  # remove .pdf
    if new_name in name_counts:
        name_counts[new_name] += 1
        suffix = chr(ord('a') + name_counts[new_name] - 1)
        new_name = f'{base_name}{suffix}.pdf'
    else:
        name_counts[new_name] = 1
    final_renames.append((old_path, new_name, rec))

# ── Print rename plan ──────────────────────────────────────────────────

print('\n' + '=' * 80)
print('RENAME PLAN')
print('=' * 80)

for old_path, new_name, rec in final_renames:
    old_display = old_path.replace('C:/OneDrive/github/pndr_survey/data/2-papers/', '')
    print(f'  {old_display}')
    print(f'    -> {new_name}')
    print()

print(f'\nTotal files to rename: {len(final_renames)}')

# ── Execute renames ────────────────────────────────────────────────────

print('\n' + '=' * 80)
print('EXECUTING RENAMES')
print('=' * 80)

for old_path, new_name, rec in final_renames:
    new_path = os.path.join(ROOT, new_name)
    if os.path.exists(new_path) and os.path.realpath(old_path) != os.path.realpath(new_path):
        print(f'  SKIP (target exists): {new_name}')
        continue
    try:
        shutil.move(old_path, new_path)
        print(f'  OK: {new_name}')
    except Exception as e:
        print(f'  ERROR: {new_name} -> {e}')

# ── Update xlsx ────────────────────────────────────────────────────────

print('\n' + '=' * 80)
print('UPDATING XLSX')
print('=' * 80)

# Build reverse mapping: (subpasta, old_filename) -> new_filename
old_to_new = {}
for old_path, new_name, rec in final_renames:
    old_fname = os.path.basename(old_path)
    old_subpasta = ''
    if '2-5-papers-anpec' in old_path:
        old_subpasta = '2-5-papers-anpec'
    elif '2-4-papers-econpapers' in old_path:
        old_subpasta = '2-4-papers-econpapers'
    old_to_new[(old_subpasta, old_fname)] = new_name

updated = 0
for row in range(2, ws.max_row + 1):
    pdf = ws.cell(row=row, column=13).value or ''
    subpasta = ws.cell(row=row, column=14).value or ''
    if not pdf:
        continue
    key = (subpasta, pdf)
    if key in old_to_new:
        ws.cell(row=row, column=13, value=old_to_new[key])
        ws.cell(row=row, column=14, value='')  # No more subfolders
        updated += 1

# Also update Arquivos Extra sheet
if 'Arquivos Extra' in wb.sheetnames:
    ws_extra = wb['Arquivos Extra']
    for row in range(2, ws_extra.max_row + 1):
        subp = ws_extra.cell(row=row, column=1).value or ''
        fname = ws_extra.cell(row=row, column=2).value or ''
        key = (subp, fname)
        if key in old_to_new:
            ws_extra.cell(row=row, column=2, value=old_to_new[key])
            ws_extra.cell(row=row, column=1, value='')

wb.save(XLSX)
print(f'Updated {updated} rows in xlsx')

# ── Clean up empty subdirs ────────────────────────────────────────────

for subdir_name in ['2-4-papers-econpapers', '2-5-papers-anpec']:
    subdir_path = os.path.join(ROOT, subdir_name)
    if os.path.isdir(subdir_path):
        remaining = os.listdir(subdir_path)
        if not remaining:
            os.rmdir(subdir_path)
            print(f'Removed empty directory: {subdir_name}')
        else:
            print(f'Directory not empty ({len(remaining)} files left): {subdir_name}')

print('\nDONE!')
