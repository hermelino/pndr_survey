"""Enriquece all_papers_llm_classif_final.xlsx com dados da etapa de extração
de registros (bib_records) e do índice de citação.

Fonte dos dados enriquecidos:
  - data/2-papers/2-2-papers.json  (merge de bib_records + all_papers + LLM)
  - data/3-ref-bib/citation_index_results.json  (índice de citação)

Matching: por campo "Arquivo PDF" no Excel → "arquivo_pdf" no JSON.
"""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

PAPERS_DIR = Path(__file__).resolve().parent.parent  # data/2-papers
XLSX = PAPERS_DIR / "all_papers_llm_classif_final.xlsx"
PAPERS_JSON = PAPERS_DIR / "2-2-papers.json"
CITATION_JSON = PAPERS_DIR.parent / "3-ref-bib" / "citation_index_results.json"

# Colunas a adicionar (label → path no JSON)
ENRICHED_COLS = [
    ("Resumo", "resumo"),
    ("Palavras-chave", "palavras_chave"),
    ("URL", "url"),
    ("ID Registro", "id_registro"),
    ("Tipo Publicação", "tipo_publicacao"),
    ("Volume", "volume"),
    ("Issue", "issue"),
    ("Pages", "pages"),
    ("Idioma", "idioma"),
]

CITATION_COLS = [
    ("Publicado?", "is_published"),
    ("IC", "IC_published"),
    ("Citações (pub.)", "citations_received_from_published"),
    ("Citações (total)", "citations_received_from_all"),
]

# XML-safe
_ILLEGAL_XML_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]"
)


def xml_safe(s):
    if not s:
        return s
    return _ILLEGAL_XML_RE.sub("", str(s))


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")


def main():
    # --- Load enriched data ---
    with open(PAPERS_JSON, "r", encoding="utf-8") as f:
        papers = json.load(f)
    papers_by_pdf = {p["arquivo_pdf"]: p for p in papers}

    with open(CITATION_JSON, "r", encoding="utf-8") as f:
        citations = json.load(f)
    # key = PDF stem (sem .pdf)
    cit_by_key = {c["key"]: c for c in citations}

    # --- Open Excel ---
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Classificação LLM"]
    headers = [cell.value for cell in ws[1]]

    # Encontrar coluna inicial para novas colunas
    # Remover colunas enriquecidas já existentes (para re-execução idempotente)
    all_new_labels = [label for label, _ in ENRICHED_COLS + CITATION_COLS]
    existing_enriched = {h: i for i, h in enumerate(headers) if h in all_new_labels}

    if existing_enriched:
        # Já existem — sobrescrever nas mesmas posições
        start_col = min(existing_enriched.values()) + 1  # 1-indexed
        print(f"  Colunas enriquecidas já existem (col {start_col}+), sobrescrevendo...")
    else:
        start_col = len(headers) + 1
        print(f"  Adicionando {len(all_new_labels)} novas colunas a partir da col {start_col}...")

    # --- Write headers ---
    for i, label in enumerate(all_new_labels):
        col_idx = start_col + i
        c = ws.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # --- Locate PDF column ---
    pdf_col_idx = headers.index("Arquivo PDF")

    # --- Write data ---
    matched = 0
    cit_matched = 0

    for row_idx in range(2, ws.max_row + 1):
        pdf_name = ws.cell(row=row_idx, column=pdf_col_idx + 1).value
        if not pdf_name:
            continue

        pdf_name = str(pdf_name).strip()
        paper = papers_by_pdf.get(pdf_name, {})

        # Enriched cols from 2-2-papers.json
        for i, (label, field) in enumerate(ENRICHED_COLS):
            col_idx = start_col + i
            val = paper.get(field, "")
            if isinstance(val, bool):
                val = "Sim" if val else "Não"
            ws.cell(row=row_idx, column=col_idx, value=xml_safe(val))
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        if paper:
            matched += 1

        # Citation cols from citation_index_results.json
        pdf_stem = Path(pdf_name).stem
        cit = cit_by_key.get(pdf_stem, {})

        offset = len(ENRICHED_COLS)
        for i, (label, field) in enumerate(CITATION_COLS):
            col_idx = start_col + offset + i
            val = cit.get(field, "")
            if isinstance(val, bool):
                val = "Sim" if val else "Não"
            elif isinstance(val, float):
                val = round(val, 4)
            ws.cell(row=row_idx, column=col_idx, value=xml_safe(val) if isinstance(val, str) else val)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="center"
            )

        if cit:
            cit_matched += 1

    # --- Column widths ---
    col_widths = {
        "Resumo": 60,
        "Palavras-chave": 35,
        "URL": 40,
        "ID Registro": 30,
        "Tipo Publicação": 20,
        "Volume": 8,
        "Issue": 8,
        "Pages": 12,
        "Idioma": 10,
        "Publicado?": 12,
        "IC": 8,
        "Citações (pub.)": 14,
        "Citações (total)": 14,
    }
    for i, label in enumerate(all_new_labels):
        col_letter = openpyxl.utils.get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = col_widths.get(label, 15)

    wb.save(XLSX)
    total = ws.max_row - 1
    print(f"\nEnriquecimento concluído!")
    print(f"  {total} registros no Excel")
    print(f"  {matched}/{total} com dados de bib_records")
    print(f"  {cit_matched}/{total} com índice de citação")


if __name__ == "__main__":
    main()
