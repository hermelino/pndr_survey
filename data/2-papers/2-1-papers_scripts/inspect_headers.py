import openpyxl

filepath = r"c:\OneDrive\github\pndr_survey\data\2-papers\all_papers.xlsx"
wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
ws = wb["Registros"]

# Print header row with column indices
for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
    for cell in row:
        print(f"  Col {cell.column:>2} ({cell.coordinate}): {cell.value!r}")

print()
print("--- First Scopus data row ---")
for row in ws.iter_rows(min_row=2, values_only=False):
    base = row[0].value
    if base and str(base).strip().lower() == "scopus":
        for cell in row:
            val = str(cell.value or "")[:80]
            print(f"  Col {cell.column:>2} ({cell.coordinate}): {val!r}")
        break

wb.close()
