"""Atualiza a aba 'Resumo' de all_papers_llm_classif_final.xlsx
com estatísticas agregadas de todas as colunas S1, S2 e S3."""

import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

XLSX = Path(__file__).resolve().parent.parent / "all_papers_llm_classif_final.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _section_header(ws, row, col1, col2):
    for idx, val in enumerate([col1, col2], 1):
        c = ws.cell(row=row, column=idx, value=val)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")


def _data_row(ws, row, col1, col2, bold=False, fill=None):
    c1 = ws.cell(row=row, column=1, value=col1)
    c2 = ws.cell(row=row, column=2, value=col2)
    for c in [c1, c2]:
        c.border = THIN_BORDER
    c2.alignment = Alignment(horizontal="center")
    if bold:
        c1.font = Font(bold=True)
        c2.font = Font(bold=True)
    if fill:
        c1.fill = fill
        c2.fill = fill


def _write_counter(ws, row, title1, title2, counter, total_val=None, total_label="Total"):
    _section_header(ws, row, title1, title2)
    row += 1
    for k, v in counter.most_common():
        _data_row(ws, row, k, v)
        row += 1
    if total_val is not None:
        _data_row(ws, row, total_label, total_val, bold=True, fill=LIGHT_FILL)
        row += 1
    return row + 1  # blank row


def _normalize(val):
    """Remove page refs like [p. X], [impl.], [ne] and capitalize."""
    if val is None or str(val).strip() in ("", "[ne]"):
        return None
    s = str(val).strip()
    s = re.sub(r"\s*\[(?:p\.[\s\d,;]+|impl\.|ne)\]$", "", s).strip()
    s = re.sub(r"\s*\[p\..*?\]$", "", s).strip()
    if s:
        s = s[0].upper() + s[1:]
    return s or None


def _bucket(val, rules):
    """Match val.lower() against rules dict {substring: label}."""
    if val is None:
        return None
    vl = val.lower()
    for substr, label in rules.items():
        if substr in vl:
            return label
    return val[:60]


# ── Categorização de variáveis dependentes ────────────────────────────────
VAR_DEP_CATEGORIES = [
    ("PIB/Renda", [
        "pib", "gdp", "renda", "income", "produto", "output", "gva",
        "valor adicionado", "consumo", "consumption", "investimento",
        "investment", "atividade", "activity level",
    ]),
    ("Emprego", [
        "emprego", "employment", "vínculo", "vinculo", "ocupação", "ocupacao",
        "trabalhador", "job", "labor", "trabalho formal",
    ]),
    ("Salários/Massa salarial", [
        "salário", "salario", "wage", "massa salarial",
        "remuneração", "remuneracao", "rendimento",
    ]),
    ("Produtividade", ["produtividade", "productivity"]),
    ("Produção agropecuária", [
        "agrícola", "agricola", "agropecuária", "agropecuaria",
        "pecuária", "pecuaria", "livestock", "vbp", "rebanho",
        "produção bruta", "agricultural",
    ]),
    ("Desigualdade/Pobreza", [
        "gini", "pobreza", "poverty", "bolsa família", "bolsa familia",
        "informalidade", "informality", "desigualdade", "inequality",
    ]),
    ("Desenvolvimento social", [
        "ifdm", "mortalidade", "mortality", "educação", "educacao",
        "ideb", "saeb", "saúde", "saude", "health", "pré-natal",
        "pre-natal", "prenatal", "nascido", "birth",
    ]),
    ("Receitas municipais", [
        "receita", "revenue", "arrecadação", "arrecadacao", "icms",
        "tribut", "fiscal", "transferência", "transferencia",
    ]),
    ("Sobrevivência de firmas", [
        "sobrevivência", "sobrevivencia", "survival", "fechamento",
        "closure", "cnpj",
    ]),
]
CATEGORY_NAMES = [c[0] for c in VAR_DEP_CATEGORIES]

# ── Normalização de nomes de instrumentos ─────────────────────────────────
INSTRUMENT_NORM = {
    "fne": "FNE", "fno": "FNO", "fco": "FCO",
    "fdne": "FDNE", "fda": "FDA", "fdco": "FDCO",
    "incentivos fiscais sudene": "IF – Sudene",
    "incentivos fiscais da sudene": "IF – Sudene",
    "incentivos fiscais sudam": "IF – Sudam",
    "incentivos fiscais da sudam": "IF – Sudam",
    "if-sudene": "IF – Sudene", "if – sudene": "IF – Sudene",
    "if-sudam": "IF – Sudam", "if – sudam": "IF – Sudam",
}
INSTRUMENT_ORDER = [
    "FNE", "FNO", "FCO", "FDNE", "FDA", "FDCO",
    "IF – Sudene", "IF – Sudam",
]
DIR_LABELS = ["Positivo", "Negativo", "Misto", "Nulo", "Não informado"]


def _normalize_instrument(name):
    """Normaliza nome do instrumento para forma canônica.

    Retorna None para instrumentos fora do escopo da PNDR
    (ex.: BNDES, outros, Prodepe).
    """
    s = name.strip().lower().rstrip(")")
    if s in INSTRUMENT_NORM:
        return INSTRUMENT_NORM[s]
    for key, val in INSTRUMENT_NORM.items():
        if key in s:
            return val
    return None  # ignora instrumentos fora do escopo


def _categorize_var_dep(var_dep_str):
    """Retorna set de categorias que casam com a string de variável dependente."""
    if not var_dep_str:
        return set()
    vl = var_dep_str.lower()
    cats = set()
    for cat_name, keywords in VAR_DEP_CATEGORIES:
        for kw in keywords:
            if kw in vl:
                cats.add(cat_name)
                break
    return cats if cats else {"Outros"}


def _section_header_wide(ws, row, first_col, col_names):
    """Escreve cabeçalho de uma matriz com múltiplas colunas."""
    c = ws.cell(row=row, column=1, value=first_col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center")
    for i, name in enumerate(col_names, 2):
        c = ws.cell(row=row, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _data_row_wide(ws, row, first_col, values, bold=False, fill=None):
    """Escreve linha de dados de uma matriz com múltiplas colunas."""
    c1 = ws.cell(row=row, column=1, value=first_col)
    c1.border = THIN_BORDER
    if bold:
        c1.font = Font(bold=True)
    if fill:
        c1.fill = fill
    for i, val in enumerate(values, 2):
        c = ws.cell(row=row, column=i, value=val if val else "")
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
        if bold:
            c.font = Font(bold=True)
        if fill:
            c.fill = fill


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws_main = wb["Classificação LLM"]
    headers = [cell.value for cell in ws_main[1]]

    col = {h: i for i, h in enumerate(headers) if h}
    triagem_idx = col["Triagem"]
    base_idx = col["Base"]
    ano_idx = col["Ano"]
    motivo_idx = next((v for k, v in col.items() if "motivo" in k.lower()), None)

    # Counters
    triagem_c = Counter()
    base_c = Counter()
    ano_c = Counter()
    motivo_c = Counter()
    tipo_trabalho_c = Counter()
    tipo_pub_c = Counter()       # Artigo / Congresso / Texto p/ discussão
    revista_c = Counter()        # Revistas dos artigos publicados
    instrumentos_c = Counter()
    metodo_c = Counter()
    tipo_dados_c = Counter()
    setor_c = Counter()
    unidade_espacial_c = Counter()
    direcao_c = Counter()
    significancia_c = Counter()
    inst_x_cat = defaultdict(Counter)     # instrumento → {categoria: n}
    inst_direction = defaultdict(Counter)  # instrumento → {direção: n}

    tipo_dados_rules = {
        "painel": "Painel",
        "mista": "Mista",
        "cross": "Cross-section",
        "serie": "Série temporal",
        "temporal": "Série temporal",
    }
    setor_rules = {
        "não especificado": "Não especificado",
        "nao especificado": "Não especificado",
        "todos": "Todos os setores",
        "agricultura": "Agricultura/Agropecuária",
        "agro": "Agricultura/Agropecuária",
        "indústria": "Indústria",
        "industria": "Indústria",
    }
    unidade_rules = {
        "munic": "Município",
        "estado": "Estado/UF",
        "uf": "Estado/UF",
        "firma": "Firma/Empresa",
        "empresa": "Firma/Empresa",
        "regi": "Região",
    }
    direcao_rules = {
        "positivo": "Positivo",
        "negativo": "Negativo",
        "misto": "Misto",
        "nulo": "Nulo",
        "não informado": "Não informado",
    }
    sig_rules = {
        "parcial": "Parcialmente",
        "não informado": "Não informado",
    }

    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        t = row[triagem_idx].value
        triagem_c[t] += 1

        if t == "APROVADO":
            base_c[row[base_idx].value] += 1
            ano_c[row[ano_idx].value] += 1

            v = _normalize(row[col.get("S1_tipo_trabalho", 0)].value)
            if v:
                tipo_trabalho_c[v] += 1

            # Tipo de publicação (bucket) e revista
            if v:
                vl = v.lower()
                if "artigo" in vl:
                    tipo_pub_c["Artigo publicado em periódico"] += 1
                    # Revista: prefer Periodico (mais padronizado), fallback S1_revista
                    rev = _normalize(row[col.get("Periodico", 0)].value)
                    if not rev:
                        rev = _normalize(row[col.get("S1_revista", 0)].value)
                    if rev:
                        revista_c[rev] += 1
                    else:
                        revista_c["(não identificada)"] += 1
                elif "congresso" in vl or "apresenta" in vl:
                    tipo_pub_c["Apresentação em congresso"] += 1
                elif "discuss" in vl:
                    tipo_pub_c["Texto para discussão"] += 1
                elif "outro" in vl:
                    tipo_pub_c["Outro"] += 1
                else:
                    tipo_pub_c[v] += 1

            # Instrumentos: split by comma + tabulação cruzada
            v_raw = row[col.get("S1_instrumentos_pndr", 0)].value
            if v_raw and str(v_raw).strip() not in ("", "[ne]"):
                cleaned = re.sub(r"\s*\[.*?\]", "", str(v_raw))

                # Variável dependente e direção para tabulação cruzada
                vd_idx = col.get("S2_var_dependente")
                var_dep = _normalize(row[vd_idx].value) if vd_idx else None
                cats = _categorize_var_dep(var_dep)

                de_idx = col.get("S3_direcao_efeito")
                dir_raw = _normalize(row[de_idx].value) if de_idx else None
                dir_norm = _bucket(dir_raw, direcao_rules) or "Não informado"

                for inst in cleaned.split(","):
                    inst = inst.strip()
                    if inst:
                        norm = _normalize_instrument(inst)
                        if norm is None:
                            continue
                        instrumentos_c[norm] += 1
                        for cat in cats:
                            inst_x_cat[norm][cat] += 1
                        inst_direction[norm][dir_norm] += 1

            v = _normalize(row[col.get("S2_metodo_econometrico", 0)].value)
            if v:
                metodo_c[v] += 1

            v = _normalize(row[col.get("S2_tipo_dados", 0)].value)
            b = _bucket(v, tipo_dados_rules)
            if b:
                tipo_dados_c[b] += 1

            v = _normalize(row[col.get("S2_setor_economico", 0)].value)
            b = _bucket(v, setor_rules)
            if b:
                setor_c[b] += 1

            v = _normalize(row[col.get("S2_unidade_espacial", 0)].value)
            b = _bucket(v, unidade_rules)
            if b:
                unidade_espacial_c[b] += 1

            v = _normalize(row[col.get("S3_direcao_efeito", 0)].value)
            b = _bucket(v, direcao_rules)
            if b:
                direcao_c[b] += 1

            v = _normalize(row[col.get("S3_significancia", 0)].value)
            if v:
                vl = v.lower()
                b = _bucket(v, sig_rules)
                if b and b != v:
                    significancia_c[b] += 1
                elif vl.startswith("sim"):
                    significancia_c["Sim"] += 1
                elif vl.startswith("não") or vl.startswith("nao"):
                    significancia_c["Não"] += 1
                else:
                    significancia_c[v] += 1

        elif t == "REJEITADO" and motivo_idx is not None:
            m = row[motivo_idx].value
            if m:
                motivo_c[str(m).strip()] += 1

    total = ws_main.max_row - 1
    aprovados = triagem_c.get("APROVADO", 0)
    rejeitados = triagem_c.get("REJEITADO", 0)

    # ── Build Resumo sheet ──────────────────────────────────────────────────
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]
    ws = wb.create_sheet("Resumo")

    r = 1

    # 1. Visão Geral
    _section_header(ws, r, "Visão Geral", "Valor"); r += 1
    _data_row(ws, r, "Total de papers analisados", total); r += 1
    _data_row(ws, r, "Aprovados na triagem", aprovados, bold=True); r += 1
    _data_row(ws, r, "Rejeitados na triagem", rejeitados); r += 1
    _data_row(ws, r, "Erros de processamento", 0); r += 1
    r += 1

    # 2. Motivos de Exclusão
    r = _write_counter(ws, r, "Motivo de Exclusão", "Qtd", motivo_c, rejeitados, "Total rejeitados")

    # 3. Aprovados por Base
    r = _write_counter(ws, r, "Aprovados por Base", "Qtd", base_c, aprovados)

    # 4. Aprovados por Período (faixas conforme metodo.tex)
    PERIODOS = [
        ("2005–2010", range(2005, 2011)),
        ("2011–2015", range(2011, 2016)),
        ("2016–2020", range(2016, 2021)),
        ("2021–2025", range(2021, 2026)),
    ]
    _section_header(ws, r, "Aprovados por Período", "Qtd"); r += 1
    for label, years in PERIODOS:
        qtd = sum(ano_c.get(y, 0) + ano_c.get(str(y), 0) for y in years)
        _data_row(ws, r, label, qtd); r += 1
    _data_row(ws, r, "Total", aprovados, bold=True, fill=LIGHT_FILL); r += 1
    r += 1

    # 5. Tipo de Publicação (agrupado)
    r = _write_counter(ws, r, "Tipo de Publicação (aprovados)", "Qtd", tipo_pub_c, aprovados)

    # 6. Revistas (artigos publicados)
    _section_header(ws, r, "Periódico / Revista (artigos)", "Qtd"); r += 1
    for k, v in revista_c.most_common():
        _data_row(ws, r, k, v); r += 1
    _data_row(ws, r, "Total artigos", tipo_pub_c.get("Artigo publicado em periódico", 0), bold=True, fill=LIGHT_FILL); r += 1
    r += 1

    # 7a. Instrumentos da PNDR (contagem simples)
    r = _write_counter(ws, r, "Instrumentos da PNDR (estudos)", "Qtd", instrumentos_c)

    # 7b. Instrumentos × Desfecho socioeconômico avaliado
    used_cats = [c for c in CATEGORY_NAMES
                 if any(inst_x_cat[inst][c] > 0 for inst in INSTRUMENT_ORDER)]
    if any(inst_x_cat[inst].get("Outros", 0) > 0 for inst in INSTRUMENT_ORDER):
        used_cats.append("Outros")
    _section_header_wide(ws, r, "Instrumento", used_cats + ["Total"])
    r += 1
    for inst in INSTRUMENT_ORDER:
        if instrumentos_c.get(inst, 0) > 0:
            values = [inst_x_cat[inst].get(c, 0) for c in used_cats]
            _data_row_wide(ws, r, inst, values + [instrumentos_c[inst]])
            r += 1
    # Linha de totais por categoria
    cat_totals = [sum(inst_x_cat[inst].get(c, 0) for inst in INSTRUMENT_ORDER)
                  for c in used_cats]
    _data_row_wide(ws, r, "Total", cat_totals + [sum(instrumentos_c.values())],
                   bold=True, fill=LIGHT_FILL)
    r += 2

    # 7c. Direção do efeito por instrumento
    _section_header_wide(ws, r, "Instrumento", DIR_LABELS + ["Total"])
    r += 1
    for inst in INSTRUMENT_ORDER:
        if instrumentos_c.get(inst, 0) > 0:
            values = [inst_direction[inst].get(d, 0) for d in DIR_LABELS]
            _data_row_wide(ws, r, inst, values + [instrumentos_c[inst]])
            r += 1
    r += 1

    # 8. Método Econométrico
    _section_header(ws, r, "Método Econométrico (aprovados)", "Qtd"); r += 1
    for k, v in metodo_c.most_common(25):
        _data_row(ws, r, k[:80], v); r += 1
    r += 1

    # 8. Tipo de Dados
    r = _write_counter(ws, r, "Tipo de Dados (aprovados)", "Qtd", tipo_dados_c)

    # 9. Setor Econômico
    r = _write_counter(ws, r, "Setor Econômico (aprovados)", "Qtd", setor_c)

    # 10. Unidade Espacial
    r = _write_counter(ws, r, "Unidade Espacial (aprovados)", "Qtd", unidade_espacial_c)

    # 11. Direção do Efeito
    r = _write_counter(ws, r, "Direção do Efeito (S3)", "Qtd", direcao_c)

    # 12. Significância Estatística
    r = _write_counter(ws, r, "Significância Estatística (S3)", "Qtd", significancia_c)

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    for letter in "CDEFGHIJKL":
        ws.column_dimensions[letter].width = 16

    wb.save(XLSX)
    print(f"Aba Resumo atualizada com sucesso! ({r} linhas)")
    print(f"  Total: {total} | Aprovados: {aprovados} | Rejeitados: {rejeitados}")


if __name__ == "__main__":
    main()
