"""Build verified all_records.xlsx with PDF matching based on content extraction.

Cross-base matching: all PDFs are searched against ALL records regardless
of which database the record came from.
"""
import json, re, unicodedata, hashlib, os
import openpyxl
from openpyxl.styles import Font

# XML-safe string: strip NULL bytes and control chars
_ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]'
)


def xml_safe(s):
    if not s:
        return s
    return _ILLEGAL_XML_RE.sub('', s)


# =====================================================================
# Load all data
# =====================================================================
src_path = 'C:/OneDrive/github/pndr_survey/data/1-records/all_records.xlsx'
dst_path = 'C:/OneDrive/github/pndr_survey/data/2-papers/all_records.xlsx'

wb = openpyxl.load_workbook(src_path)
ws = wb['Registros']

with open('C:/OneDrive/github/pndr_survey/data/2-papers/econpapers_text.json', 'r', encoding='utf-8') as f:
    econ_text = json.load(f)
with open('C:/OneDrive/github/pndr_survey/data/2-papers/anpec_text.json', 'r', encoding='utf-8') as f:
    anpec_text = json.load(f)

# Merge all PDFs into single pool: {filename: {text, subdir}}
all_pdfs = {}
for fname, info in econ_text.items():
    if not info.get('error'):
        all_pdfs[fname] = {'text': info['text'], 'subdir': '2-4-papers-econpapers'}
for fname, info in anpec_text.items():
    if not info.get('error'):
        all_pdfs[fname] = {'text': info['text'], 'subdir': '2-5-papers-anpec'}


# =====================================================================
# Matching utilities
# =====================================================================
def normalize(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s.lower()).strip()


def get_surnames(authors_str):
    if not authors_str:
        return []
    surnames = []
    # Handle both "; " and ", " as separators between authors
    for part in re.split(r';\s*', authors_str):
        part = part.strip()
        if ',' in part:
            surnames.append(part.split(',')[0].strip())
        elif part.split():
            surnames.append(part.split()[-1])
    return [normalize(s) for s in surnames if len(s) > 2]


def surname_in_text(surname, norm_text):
    """Check if a surname appears in text, tolerating footnote digits (e.g. 'resende2')."""
    # Try exact word boundary first
    if re.search(r'\b' + re.escape(surname) + r'\b', norm_text):
        return True
    # Allow trailing digits (footnote markers stuck to name: "resende2")
    if re.search(r'\b' + re.escape(surname) + r'\d', norm_text):
        return True
    # For compound surnames like "mendes resende", also try just the last part
    parts = surname.split()
    if len(parts) > 1:
        last = parts[-1]
        if len(last) >= 4 and re.search(r'\b' + re.escape(last) + r'[\b\d]', norm_text):
            return True
    return False


def verify_match(title, authors, pdf_text):
    """Return (title_pct, is_match) checking title words + author surnames."""
    norm_text = normalize(pdf_text)
    norm_title = normalize(title)
    words = [w for w in re.sub(r'[^a-z0-9 ]', ' ', norm_title).split() if len(w) >= 5]
    if not words:
        return 0, False
    norm_text_alpha = re.sub(r'[^a-z0-9 ]', ' ', norm_text)
    title_hits = sum(1 for w in words if w in norm_text_alpha)
    title_pct = title_hits / len(words) * 100
    surnames = get_surnames(authors)
    if surnames:
        author_hits = sum(1 for s in surnames if surname_in_text(s, norm_text))
        author_ok = author_hits >= 1
    else:
        author_ok = True
    return title_pct, title_pct >= 60 and author_ok


# =====================================================================
# Load all records
# =====================================================================
records = []
for row_idx in range(2, ws.max_row + 1):
    records.append({
        'row': row_idx,
        'base': ws.cell(row=row_idx, column=1).value,
        'id': ws.cell(row=row_idx, column=2).value or '',
        'title': ws.cell(row=row_idx, column=3).value or '',
        'authors': ws.cell(row=row_idx, column=4).value or '',
        'year': ws.cell(row=row_idx, column=5).value,
        'doi': ws.cell(row=row_idx, column=7).value or '',
        'url': ws.cell(row=row_idx, column=8).value or '',
    })


# =====================================================================
# PHASE 1: ANPEC hash-based matching (deterministic, verified)
# =====================================================================
anpec_dir = 'C:/OneDrive/github/pndr_survey/data/2-papers/2-5-papers-anpec'
anpec_files_set = set(os.listdir(anpec_dir))

# result: row -> (status, filename, subdir, note)
matches = {}

for rec in records:
    if rec['base'] == 'anpec':
        match_url = rec['url'] or rec['id']
        if match_url:
            h = hashlib.md5(match_url.encode()).hexdigest()[:6]
            yr = str(rec['year']) if rec['year'] else 'nd'
            fname = f'anpec_{yr}_{h}.pdf'
            if fname in anpec_files_set:
                matches[rec['row']] = ('Sim', fname, '2-5-papers-anpec', '')
                continue
    # Not matched yet
    if rec['row'] not in matches:
        matches[rec['row']] = None  # placeholder


# =====================================================================
# PHASE 2: URL/DOI-based deterministic matching for econpapers
# =====================================================================
for rec in records:
    if matches.get(rec['row']) is not None:
        continue  # already matched
    if rec['base'] != 'econpapers':
        continue

    url = rec['url']
    candidate = None

    m = re.search(r'ipetds:(\d+)', url)
    if m:
        td = m.group(1)
        candidates = [f for f in econ_text if f.lower() == f'td_{td}.pdf']
        if candidates:
            candidate = candidates[0]

    if not candidate:
        m = re.search(r'ufb:wpaper:(\d+)', url)
        if m:
            candidates = [f for f in econ_text if f.upper() == f'TD_{m.group(1)}.PDF']
            if candidates:
                candidate = candidates[0]

    if not candidate:
        m = re.search(r'10\.1007[_/](s[\d-]+)', url)
        if m:
            candidates = [f for f in econ_text if m.group(1) in f]
            if candidates:
                candidate = candidates[0]

    if not candidate and 'bla:jregsc' in url:
        candidates = [f for f in econ_text if 'Journal of Regional Science' in f]
        if candidates:
            candidate = candidates[0]

    if candidate and candidate in all_pdfs:
        pct, ok = verify_match(rec['title'], rec['authors'], all_pdfs[candidate]['text'])
        if ok:
            matches[rec['row']] = ('Sim', candidate, '2-4-papers-econpapers', '')


# =====================================================================
# PHASE 3: Cross-base fuzzy matching (ALL records x ALL PDFs)
# For every unmatched record, search the entire PDF pool.
# Stricter threshold for cross-base: require >=80% title AND >=2 authors
# (or >=1 if the record has only 1 author).
# =====================================================================
def verify_strict(title, authors, pdf_text):
    """Stricter verification for cross-base matching."""
    norm_text = normalize(pdf_text)
    norm_title = normalize(title)
    words = [w for w in re.sub(r'[^a-z0-9 ]', ' ', norm_title).split() if len(w) >= 5]
    if not words:
        return 0, False
    norm_text_alpha = re.sub(r'[^a-z0-9 ]', ' ', norm_text)
    title_hits = sum(1 for w in words if w in norm_text_alpha)
    title_pct = title_hits / len(words) * 100
    surnames = get_surnames(authors)
    if surnames:
        author_hits = sum(1 for s in surnames if surname_in_text(s, norm_text))
        # Require at least 2 authors (or all if only 1)
        min_needed = min(2, len(surnames))
        author_ok = author_hits >= min_needed
    else:
        author_ok = True
    return title_pct, title_pct >= 80 and author_ok


# =====================================================================
# PHASE 2.5: DOI-based matching for records with DOIs
# Searches for the DOI string inside the PDF text
# =====================================================================
for rec in records:
    if matches.get(rec['row']) is not None:
        continue
    doi = rec['doi']
    if not doi:
        continue
    # Normalize DOI for searching (lowercase, no URL prefix)
    doi_clean = doi.lower().strip()
    if doi_clean.startswith('http'):
        doi_clean = doi_clean.split('doi.org/')[-1]
    # Remove common DOI prefixes for partial matching
    doi_short = doi_clean.replace('10.', '')

    best_file = None
    best_subdir = ''
    for fname, info in all_pdfs.items():
        norm_text = normalize(info['text'])
        if doi_clean in norm_text or doi_short in norm_text:
            best_file = fname
            best_subdir = info['subdir']
            break  # DOI is unique, first match is sufficient

    if best_file:
        matches[rec['row']] = ('Sim', best_file, best_subdir, 'DOI match')


# Map record base to expected subdir for "same-base" detection
BASE_TO_SUBDIR = {
    'econpapers': '2-4-papers-econpapers',
    'anpec': '2-5-papers-anpec',
}

for rec in records:
    if matches.get(rec['row']) is not None:
        continue

    best_file = None
    best_pct = 0
    best_subdir = ''
    own_subdir = BASE_TO_SUBDIR.get(rec['base'], '')

    for fname, info in all_pdfs.items():
        is_same_base = (info['subdir'] == own_subdir) if own_subdir else False
        # Use normal threshold for same-base, strict for cross-base
        if is_same_base:
            pct, ok = verify_match(rec['title'], rec['authors'], info['text'])
        else:
            pct, ok = verify_strict(rec['title'], rec['authors'], info['text'])
        if ok and pct > best_pct:
            best_pct = pct
            best_file = fname
            best_subdir = info['subdir']

    if best_file:
        matches[rec['row']] = ('Sim', best_file, best_subdir, '')
    else:
        matches[rec['row']] = ('Nao', '', '', '')


# =====================================================================
# Write updated xlsx
# =====================================================================
ws.cell(row=1, column=12, value='Baixado').font = Font(bold=True)
ws.cell(row=1, column=13, value='Arquivo PDF').font = Font(bold=True)
ws.cell(row=1, column=14, value='Subpasta').font = Font(bold=True)
ws.cell(row=1, column=15, value='Obs').font = Font(bold=True)

green_font = Font(color='006400')
red_font = Font(color='8B0000')

assigned_econpapers = set()
assigned_anpec = set()

for row_idx in range(2, ws.max_row + 1):
    result = matches.get(row_idx, ('Nao', '', '', ''))
    status, fname, subdir, note = result

    ws.cell(row=row_idx, column=12, value=status)
    ws.cell(row=row_idx, column=13, value=fname)
    ws.cell(row=row_idx, column=14, value=subdir)
    ws.cell(row=row_idx, column=15, value=note)

    if status == 'Sim':
        ws.cell(row=row_idx, column=12).font = green_font
        if subdir == '2-4-papers-econpapers':
            assigned_econpapers.add(fname)
        elif subdir == '2-5-papers-anpec':
            assigned_anpec.add(fname)
    else:
        ws.cell(row=row_idx, column=12).font = red_font


# =====================================================================
# New sheet: Arquivos Extra (unassigned files)
# =====================================================================
if 'Arquivos Extra' in wb.sheetnames:
    del wb['Arquivos Extra']
ws_extra = wb.create_sheet('Arquivos Extra')
ws_extra.cell(row=1, column=1, value='Subpasta').font = Font(bold=True)
ws_extra.cell(row=1, column=2, value='Arquivo').font = Font(bold=True)
ws_extra.cell(row=1, column=3, value='Conteudo identificado').font = Font(bold=True)
ws_extra.cell(row=1, column=4, value='Relacionado a PNDR?').font = Font(bold=True)

extra_row = 2

# Descriptions for known unassigned files
known_extras = {
    'td_1969.pdf': ('Sim (fora dos registros)', 'IPEA TD 1969: Impactos do FCO 2004-2010 (Resende, Cravo, Pires)'),
    'i6-e11ba9187a805f3395cd09be5a703597.doc': ('?', 'Arquivo .doc - nao pode ser lido automaticamente'),
}

for fname in sorted(econ_text.keys()):
    if fname not in assigned_econpapers:
        if fname in known_extras:
            pndr, desc = known_extras[fname]
        else:
            txt = econ_text[fname].get('text', '')[:200].replace('\n', ' ')
            pndr = '?'
            desc = txt

        ws_extra.cell(row=extra_row, column=1, value='2-4-papers-econpapers')
        ws_extra.cell(row=extra_row, column=2, value=xml_safe(fname))
        ws_extra.cell(row=extra_row, column=3, value=xml_safe(desc))
        ws_extra.cell(row=extra_row, column=4, value=pndr)
        extra_row += 1

# ANPEC extras
for fname in sorted(anpec_text.keys()):
    if fname not in assigned_anpec:
        txt = anpec_text[fname].get('text', '')[:200].replace('\n', ' ')
        ws_extra.cell(row=extra_row, column=1, value='2-5-papers-anpec')
        ws_extra.cell(row=extra_row, column=2, value=xml_safe(fname))
        ws_extra.cell(row=extra_row, column=3, value=xml_safe(txt[:200]))
        ws_extra.cell(row=extra_row, column=4, value='?')
        extra_row += 1


# =====================================================================
# Update Resumo sheet
# =====================================================================
ws_resumo = wb['Resumo']
ws_resumo.cell(row=1, column=5, value='PDFs verificados').font = Font(bold=True)

base_counts = {}
for row_idx in range(2, ws.max_row + 1):
    base = ws.cell(row=row_idx, column=1).value
    status = ws.cell(row=row_idx, column=12).value
    if base not in base_counts:
        base_counts[base] = 0
    if status == 'Sim':
        base_counts[base] += 1

base_map = {
    'Scopus': 'scopus',
    'SciELO': 'scielo',
    'Portal CAPES': 'capes',
    'EconPapers/RePEc': 'econpapers',
    'ANPEC': 'anpec',
}
total_dl = 0
for row_idx in range(2, ws_resumo.max_row + 1):
    label = ws_resumo.cell(row=row_idx, column=1).value
    if label in base_map:
        c = base_counts.get(base_map[label], 0)
        ws_resumo.cell(row=row_idx, column=5, value=c)
        total_dl += c
    elif label == 'TOTAL':
        ws_resumo.cell(row=row_idx, column=5, value=total_dl)

# Column widths
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 50
ws.column_dimensions['N'].width = 22
ws.column_dimensions['O'].width = 25
ws_extra.column_dimensions['A'].width = 22
ws_extra.column_dimensions['B'].width = 55
ws_extra.column_dimensions['C'].width = 70
ws_extra.column_dimensions['D'].width = 22

wb.save(dst_path)
print(f'Salvo em: {dst_path}')

# Final summary
print()
print('RESUMO FINAL (verificado por conteudo dos PDFs):')
print('=' * 55)
for base in ['scopus', 'scielo', 'capes', 'econpapers', 'anpec']:
    total = sum(1 for r in records if r['base'] == base)
    dl = base_counts.get(base, 0)
    pct = dl / total * 100 if total else 0
    print(f'  {base:15s}: {dl:3d}/{total:3d} ({pct:.0f}%)')
print(f'  {"TOTAL":15s}: {total_dl:3d}/{len(records):3d} ({total_dl / len(records) * 100:.0f}%)')
print()
print(f'Arquivos extra (nao mapeados a registros): {extra_row - 2}')

# Show per-base detail for newly matched
print()
print('DETALHES cross-base:')
for rec in records:
    result = matches.get(rec['row'])
    if result and result[0] == 'Sim' and rec['base'] in ('scopus', 'scielo', 'capes'):
        print(f"  [{rec['base']}] {rec['title'][:60]}")
        print(f"       -> {result[1]}")
