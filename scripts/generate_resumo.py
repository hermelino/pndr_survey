"""Regenera a aba Resumo em all_papers_llm_classif_final.xlsx.

Lê de:
  - data/2-papers/2-2-papers.json (fonte MASTER)
  - data/3-ref-bib/citation_index_results.json (dados de IC)

Atualiza:
  - data/2-papers/all_papers_llm_classif_final.xlsx (apenas aba Resumo)

Tabelas geradas (estrutura alinhada com o artigo):
  1. Visão Geral
  2. Motivos de Exclusão
  3. Aprovados por Período (tab:estudos-ano)
  4. Instrumentos PNDR (tab:instrumentos)
  5. Unidade Amostral (tab:unidade-amostral)
  6. Métodos Econométricos (tab:metodos) — TODOS, normalizados, com MSM
  7. Autores (tab:autores-todos) — TODOS
  8. Índice de Citação Cruzada (IC)
"""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

PAPERS_JSON = DATA_DIR / "2-papers" / "2-2-papers.json"
IC_JSON = DATA_DIR / "3-ref-bib" / "citation_index_results.json"
OUTPUT_XLSX = DATA_DIR / "2-papers" / "all_papers_llm_classif_final.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SECTION_FONT = Font(bold=True, size=11)
BOLD_FONT = Font(bold=True)

# ---------------------------------------------------------------------------
# MSM scores for each normalized method
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Author alias mapping — merges known name variants to canonical form.
# Built by comparing normalized output against the article's curated table.
# ---------------------------------------------------------------------------
AUTHOR_ALIASES: dict[str, str] = {
    # Irffi, G. D. (article: 11)
    "Irffi, G.": "Irffi, G. D.",
    # Resende, G. M. (article: 10)
    "Resende, G.": "Resende, G. M.",
    "Mendes Resende, G.": "Resende, G. M.",
    "Mendes Resende, G. M.": "Resende, G. M.",
    # Carneiro, D. R. F. (article: 9)
    "Carneiro, D.": "Carneiro, D. R. F.",
    # Bastos, F. S. (article: 4)
    "Bastos, F.": "Bastos, F. S.",
    "de Sousa Bastos, F.": "Bastos, F. S.",
    "de Sousa Bastos, F. S.": "Bastos, F. S.",
    # Braz, M. S. (article: 4)
    "Braz, M.": "Braz, M. S.",
    # Costa, E. M. (article: 4)
    "Costa, E.": "Costa, E. M.",
    # Silveira Neto, R. M. (article: 4)
    "Neto, R. M. S.": "Silveira Neto, R. M.",
    "Neto, R. M.": "Silveira Neto, R. M.",
    "Neto, R.": "Silveira Neto, R. M.",
    # Veloso, P. A. S. (article: 4)
    "Veloso, P.": "Veloso, P. A. S.",
    "Velooso, P. A. S.": "Veloso, P. A. S.",  # LLM typo
    # Other known short-initial variants
    "Soares, R.": "Soares, R. B.",
    "Dias, T.": "Dias, T. K.",
    "Domingues, E.": "Domingues, E. P.",
    "Nunes, E.": "Nunes, E. S.",
    # Abel da Silva Filho — LLM outputs "FILHO, L. A. S." instead of compound
    "Filho, L. A. S.": "Silva Filho, L. A.",
    "Abel da Silva Filho, L.": "Silva Filho, L. A.",
    # da Silva variants for Diego Firmino Costa da Silva
    "Silva, D. F.": "Silva, D. F. C.",
    "Firmino Costa da Silva, D.": "Silva, D. F. C.",
    # da Silva, A. M. A. vs Silva, A. M. A.
    "Silva, A. M. A.": "da Silva, A. M. A.",
    "Silva, A.": "da Silva, A. M. A.",
    # LLM typos
    "Gondiim, I.": "Gondim, I.",
}

MSM_SCORES = {
    "Diferenças em Diferenças Escalonado": "3",
    "Diferenças em Diferenças (DiD)": "3",
    "Propensity Score Matching (PSM)": "3",
    "Generalized Propensity Score (GPS)": "3",
    "Análise Envoltória de Dados (DEA)": "n.c.",
    "Equilíbrio Geral Computável (EGC)": "n.c.",
    "Controle Sintético Generalizado": "3",
    "Regressão Descontínua (RDD)": "4",
    "Variáveis Instrumentais (IV)": "3",
    "Painel Espacial": "3",
    "Modelo de Erro Espacial": "3",
    "Painel de Efeitos Fixos": "3",
    "Painel de Efeitos Aleatórios": "3",
    "Painel Dinâmico GMM": "3",
    "MQO/OLS": "2",
    "Primeiras Diferenças (FD)": "3",
    "Fronteira Estocástica (SFA)": "n.c.",
    "Índice de Malmquist": "n.c.",
    "Fronteira de Ordem-m": "n.c.",
    "Modelo de Efeito Limiar (Threshold)": "3",
    "Análise Exploratória Espacial (AEDE)": "n.c.",
    "Análise Espacial": "n.c.",
    "Análise de Sobrevivência": "n.c.",
}


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------
def _strip_refs(s: str) -> str:
    """Remove page references like [p. 5] from a string."""
    return re.sub(r"\s*\[.*?\]", "", s).strip()


def normalize_methods(raw: str) -> list[str]:
    """Extract and normalize method names from raw LLM description.

    A single study may employ multiple methods; this function returns
    all identified methods as a list of standardized names.
    """
    if not raw or raw.strip() in ("", "[ne]", "ne"):
        return []

    raw = _strip_refs(raw)
    raw_lower = raw.lower()
    found = []

    # DiD Escalonado (check before generic DiD)
    if any(kw in raw_lower for kw in [
        "escalonado", "staggered", "callaway", "múltiplos períodos",
    ]):
        if any(kw in raw_lower for kw in ["diferenças", "diff"]):
            found.append("Diferenças em Diferenças Escalonado")

    # Generic DiD (only if escalonado not matched)
    if "Diferenças em Diferenças Escalonado" not in found:
        if any(kw in raw_lower for kw in [
            "diferenças em diferenças", "diff-in-diff",
        ]):
            found.append("Diferenças em Diferenças (DiD)")

    # RDD
    if any(kw in raw_lower for kw in [
        "regressão descontínua", "regression discontinuity", "grdd",
    ]):
        found.append("Regressão Descontínua (RDD)")

    # GPS (before PSM)
    if any(kw in raw_lower for kw in [
        "generalized propensity", "escore de propensão generalizado",
        "dose-response",
    ]):
        found.append("Generalized Propensity Score (GPS)")

    # PSM
    if any(kw in raw_lower for kw in [
        "propensity score matching", "pareamento por escore de propensão",
    ]):
        found.append("Propensity Score Matching (PSM)")

    # DEA
    if any(kw in raw_lower for kw in [
        "envoltória de dados", "data envelopment",
    ]) or re.search(r"\bdea\b", raw_lower):
        found.append("Análise Envoltória de Dados (DEA)")

    # EGC/CGE
    if any(kw in raw_lower for kw in [
        "equilíbrio geral", "general equilibrium",
    ]) or re.search(r"\b(cge|egc)\b", raw_lower):
        found.append("Equilíbrio Geral Computável (EGC)")

    # Synthetic Control
    if any(kw in raw_lower for kw in ["synthetic control", "controle sintético"]):
        found.append("Controle Sintético Generalizado")

    # IV
    if any(kw in raw_lower for kw in [
        "variáveis instrumentais", "instrumental variables",
    ]):
        found.append("Variáveis Instrumentais (IV)")

    # Spatial Panel (before generic FE)
    if any(kw in raw_lower for kw in [
        "painel espacial", "spatial durbin", "sdem",
        "econometria espacial",
    ]):
        found.append("Painel Espacial")

    # Spatial Error Model
    if any(kw in raw_lower for kw in ["erro espacial", "spatial error"]):
        found.append("Modelo de Erro Espacial")

    # FE Panel
    if any(kw in raw_lower for kw in [
        "efeito fixo", "efeitos fixos", "fixed effect",
    ]):
        found.append("Painel de Efeitos Fixos")

    # RE Panel
    if any(kw in raw_lower for kw in ["efeitos aleatórios", "random effect"]):
        found.append("Painel de Efeitos Aleatórios")

    # Dynamic Panel GMM
    if "painel dinâmico" in raw_lower or "dynamic panel" in raw_lower:
        found.append("Painel Dinâmico GMM")
    elif re.search(r"\bgmm\b", raw_lower) and "Modelo de Erro Espacial" not in found:
        found.append("Painel Dinâmico GMM")

    # OLS/MQO (after more specific methods)
    if any(kw in raw_lower for kw in ["mínimos quadrados", "mqo"]):
        found.append("MQO/OLS")
    elif re.search(r"\bols\b", raw_lower):
        found.append("MQO/OLS")

    # First Differences
    if any(kw in raw_lower for kw in ["first-differenc", "primeiras diferenças"]):
        found.append("Primeiras Diferenças (FD)")

    # SFA
    if any(kw in raw_lower for kw in [
        "fronteira estocástica", "stochastic frontier",
    ]) or re.search(r"\bsfa\b", raw_lower):
        found.append("Fronteira Estocástica (SFA)")

    # Malmquist
    if "malmquist" in raw_lower:
        found.append("Índice de Malmquist")

    # Frontier of order-m
    if any(kw in raw_lower for kw in ["fronteira de ordem", "ordem-m"]):
        found.append("Fronteira de Ordem-m")

    # Threshold
    if any(kw in raw_lower for kw in ["efeito limiar", "threshold"]):
        found.append("Modelo de Efeito Limiar (Threshold)")

    # AEDE
    if any(kw in raw_lower for kw in ["análise exploratória", "aede"]):
        found.append("Análise Exploratória Espacial (AEDE)")

    # Spatial Analysis (generic, only if no specific spatial matched)
    if "análise espacial" in raw_lower and not any(
        "Espacial" in f for f in found
    ):
        found.append("Análise Espacial")

    # Survival analysis
    if any(kw in raw_lower for kw in ["sobrevivência", "survival"]):
        found.append("Análise de Sobrevivência")

    # Fallback: show raw string if nothing matched
    if not found and raw:
        found.append(raw)

    return found


def normalize_instruments(raw: str) -> list[str]:
    """Extract and normalize PNDR instruments from raw LLM description."""
    if not raw or raw.strip() in ("", "[ne]", "ne"):
        return []

    raw = _strip_refs(raw)
    raw_upper = raw.upper()
    found = []

    PATTERNS = [
        ("FNE", [r"\bFNE\b"]),
        ("FNO", [r"\bFNO\b"]),
        ("FCO", [r"\bFCO\b"]),
        ("FDNE", [r"\bFDNE\b"]),
        ("FDA", [r"\bFDA\b"]),
        ("FDCO", [r"\bFDCO\b"]),
        ("IF -- Sudene", [
            r"IF\s*[-\u2013\u2014]\s*SUDENE",
            r"INCENTIVO\w*\s+FISCA\w*.*SUDENE",
        ]),
        ("IF -- Sudam", [
            r"IF\s*[-\u2013\u2014]\s*SUDAM",
            r"INCENTIVO\w*\s+FISCA\w*.*SUDAM",
        ]),
    ]

    for name, pats in PATTERNS:
        for pat in pats:
            if re.search(pat, raw_upper):
                found.append(name)
                break

    if not found and raw:
        found.append(raw)

    return found


def normalize_author_name(raw: str) -> str:
    """Normalize an author name to 'Sobrenome, I. I.' format.

    Handles both formats:
      - "Sobrenome, Nome Completo" (citation format)
      - "Nome Completo Sobrenome" (full-name format)

    Based on logic from temp_author_analysis.py.
    """
    SUFFIXES = {"junior", "filho", "neto", "sobrinho", "segundo", "terceiro",
                "júnior"}
    PARTICLES = {"de", "da", "do", "dos", "das", "del"}

    raw = _strip_refs(raw).strip()
    if not raw:
        return ""

    def _title_word(w: str) -> str:
        return w.lower() if w.lower() in PARTICLES else w.capitalize()

    def _initials(words: list[str]) -> list[str]:
        return [
            w[0].upper() + "."
            for w in words
            if w.lower() not in PARTICLES and w and w[0].isalpha()
        ]

    if "," in raw:
        # Citation format: "Sobrenome, Nomes"
        parts = raw.split(",", 1)
        surname_raw = parts[0].strip()
        firstnames_raw = parts[1].strip() if len(parts) > 1 else ""
        # Fix concatenated initials: "D.R.F." → "D. R. F."
        firstnames_raw = re.sub(r"\.(?=[A-Za-z])", ". ", firstnames_raw)
        surname = " ".join(_title_word(w) for w in surname_raw.split())
        if not firstnames_raw:
            return surname
        inits = _initials(firstnames_raw.split())
        return f"{surname}, {' '.join(inits)}" if inits else surname
    else:
        # Full-name format: "Nome Completo Sobrenome"
        words = raw.split()
        if not words:
            return raw
        surname_parts: list[str] = []
        idx = len(words) - 1
        # Collect suffixes (Filho, Neto, etc.)
        while idx >= 0 and words[idx].lower().replace(".", "") in SUFFIXES:
            surname_parts.insert(0, words[idx])
            idx -= 1
        # Main surname
        if idx >= 0:
            surname_parts.insert(0, words[idx])
            idx -= 1
        # Particles (de, da, etc.)
        while idx >= 0 and words[idx].lower() in PARTICLES:
            surname_parts.insert(0, words[idx])
            idx -= 1
        firstnames = words[: idx + 1]
        surname = " ".join(_title_word(w) for w in surname_parts)
        inits = _initials(firstnames)
        return f"{surname}, {' '.join(inits)}" if inits else surname


def get_all_authors(paper: dict) -> list[str]:
    """Extract and normalize ALL authors (autorias + coautorias) for a paper.

    Uses s1.autores (LLM-extracted) as **primary** source because:
      - LLM always returns full first names (better normalization)
      - LLM uses consistent "SURNAME, Firstname" semicolon-separated format
      - 18 ANPEC papers have empty bib metadata
    Falls back to bib ``autores`` when s1 is unavailable.
    """
    raw = paper.get("s1", {}).get("autores", "").strip()
    if not raw:
        raw = paper.get("autores", "").strip()
    if not raw:
        return []

    authors = []
    # s1.autores uses ";" separator: "SOBRENOME, Nome; SOBRENOME, Nome"
    # bib autores also uses ";" for most entries
    if ";" in raw:
        parts = raw.split(";")
    else:
        # Some bib entries use "," between full names (no ";")
        # e.g. "Fulano de Tal, Ciclano de Tal, Beltrano"
        # Heuristic: if no ";" but has uppercase words pattern, split by ","
        parts = raw.split(",")
        # But "Sobrenome, Nome" format also uses commas...
        # If first part looks like a surname (short, capitalized), it's citation
        # format with a single author -> treat whole string as one author
        if len(parts) == 2 and len(parts[0].split()) <= 3:
            parts = [raw]  # single author in "Surname, Firstname" format

    for part in parts:
        name = normalize_author_name(part)
        if name:
            authors.append(name)
    return authors


def normalize_unit(raw: str) -> list[str]:
    """Normalize spatial unit of analysis."""
    if not raw or raw.strip() in ("", "[ne]", "ne"):
        return []

    raw = _strip_refs(raw)
    raw_lower = raw.lower()
    found = []

    if any(kw in raw_lower for kw in ["município", "municipal", "municipio"]):
        found.append("Município")
    if any(kw in raw_lower for kw in ["empresa", "firma", "empreendimento"]):
        found.append("Empresa")
    if re.search(r"\buf\b|unidade\w* da federa", raw_lower) or "estado" in raw_lower:
        found.append("UF")
    if any(kw in raw_lower for kw in ["área mínima comparável", "amc"]):
        found.append("Área Mínima Comparável")
    if "microrregião" in raw_lower or "microrregiã" in raw_lower:
        found.append("Microrregião")
    if "mesorregião" in raw_lower or "mesorregiã" in raw_lower:
        found.append("Mesorregião")

    if not found and raw:
        found.append(raw)

    return found


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------
def write_header_row(ws, row: int, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_section_title(ws, row: int, title: str) -> None:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate_resumo() -> None:
    print("Carregando dados...")

    with open(PAPERS_JSON, encoding="utf-8") as f:
        papers = json.load(f)

    with open(IC_JSON, encoding="utf-8") as f:
        ic_data = json.load(f)

    aprovados = [p for p in papers if p.get("triagem") == "APROVADO"]
    rejeitados = [p for p in papers if p.get("triagem") == "REJEITADO"]
    n_total = len(papers)
    n_aprov = len(aprovados)
    n_rej = len(rejeitados)
    print(f"  Total: {n_total}, Aprovados: {n_aprov}, Rejeitados: {n_rej}")

    # Load workbook (touch only the Resumo sheet)
    wb = openpyxl.load_workbook(str(OUTPUT_XLSX))
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]
    ws = wb.create_sheet("Resumo")

    row = 1

    # ===== 1. Visão Geral =====
    write_section_title(ws, row, "Visão Geral")
    row += 1
    write_header_row(ws, row, ["Métrica", "Valor"])
    row += 1
    for label, val in [
        ("Total de registros (pós-dedup)", n_total),
        ("Aprovados", n_aprov),
        ("Rejeitados", n_rej),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # ===== 2. Motivos de Exclusão =====
    row += 1
    write_section_title(ws, row, "Motivos de Exclusão")
    row += 1
    write_header_row(ws, row, ["Motivo", "Qtd."])
    row += 1
    motivos = Counter()
    for p in rejeitados:
        motivo = p.get("motivo_exclusao", "") or "Não informado"
        # Strip "LLM: " prefix to group manual and LLM classifications
        motivo = re.sub(r"^LLM:\s*", "", motivo)
        motivos[motivo] += 1
    for motivo, count in motivos.most_common():
        ws.cell(row=row, column=1, value=motivo)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # ===== 3. Aprovados por Período (tab:estudos-ano) =====
    row += 1
    write_section_title(ws, row, "Aprovados por Período (tab:estudos-ano)")
    row += 1
    write_header_row(ws, row, ["Período", "Qtd. Artigos"])
    row += 1
    periodos = [
        ("2005--2010", 2005, 2010),
        ("2011--2015", 2011, 2015),
        ("2016--2020", 2016, 2020),
        ("2021--2025", 2021, 2025),
    ]
    total_periodo = 0
    for label, ini, fim in periodos:
        count = sum(1 for p in aprovados if ini <= int(p.get("ano", 0)) <= fim)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        total_periodo += count
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
    ws.cell(row=row, column=2, value=total_periodo).font = BOLD_FONT
    row += 1

    # ===== 4. Instrumentos PNDR (tab:instrumentos) =====
    row += 1
    write_section_title(ws, row, "Instrumentos PNDR (tab:instrumentos)")
    row += 1
    write_header_row(ws, row, ["Instrumento", "Qtd. Artigos"])
    row += 1
    inst_counts = Counter()
    for p in aprovados:
        raw = p.get("s1", {}).get("instrumentos_pndr", "")
        for inst in normalize_instruments(raw):
            inst_counts[inst] += 1
    # Fixed order matching article
    inst_order = [
        "FNE", "FNO", "FCO", "FDNE", "FDA", "FDCO",
        "IF -- Sudene", "IF -- Sudam",
    ]
    shown = set()
    for inst in inst_order:
        if inst in inst_counts:
            ws.cell(row=row, column=1, value=inst)
            ws.cell(row=row, column=2, value=inst_counts[inst])
            shown.add(inst)
            row += 1
    for inst, count in inst_counts.most_common():
        if inst not in shown:
            ws.cell(row=row, column=1, value=inst)
            ws.cell(row=row, column=2, value=count)
            row += 1

    # ===== 5. Unidade Amostral (tab:unidade-amostral) =====
    row += 1
    write_section_title(ws, row, "Unidade Amostral (tab:unidade-amostral)")
    row += 1
    write_header_row(ws, row, ["Unidade Amostral", "Qtd. Artigos"])
    row += 1
    unit_counts = Counter()
    for p in aprovados:
        raw = p.get("s2", {}).get("unidade_espacial", "")
        for u in normalize_unit(raw):
            unit_counts[u] += 1
    for u, count in unit_counts.most_common():
        ws.cell(row=row, column=1, value=u)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # ===== 6. Métodos Econométricos (tab:metodos) =====
    row += 1
    write_section_title(ws, row, "Métodos Econométricos (tab:metodos)")
    row += 1
    write_header_row(ws, row, ["Método", "Estudos", "MSM"])
    row += 1
    method_counts = Counter()
    for p in aprovados:
        raw = p.get("s2", {}).get("metodo_econometrico", "")
        for m in normalize_methods(raw):
            method_counts[m] += 1
    for m, count in method_counts.most_common():
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=MSM_SCORES.get(m, "?"))
        row += 1

    # ===== 7. Autores — autorias e coautorias (tab:autores-todos) =====
    row += 1
    write_section_title(ws, row, "Autores — autorias e coautorias (tab:autores-todos)")
    row += 1
    write_header_row(ws, row, ["Autor", "Qtd. Artigos"])
    row += 1
    author_counts = Counter()
    for p in aprovados:
        for name in get_all_authors(p):
            name = AUTHOR_ALIASES.get(name, name)  # apply alias dedup
            author_counts[name] += 1
    for a, count in author_counts.most_common():
        ws.cell(row=row, column=1, value=a)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # ===== 8. Índice de Citação Cruzada (IC) =====
    row += 1
    write_section_title(ws, row, "Índice de Citação Cruzada (IC)")
    row += 1
    ic_headers = ["Estudo", "Ano", "Pub?", "Cit.", "N", "IC", "Tipo"]
    write_header_row(ws, row, ic_headers)
    row += 1
    ic_sorted = sorted(
        ic_data,
        key=lambda x: (-x.get("IC_published", 0), x.get("key", "")),
    )
    for entry in ic_sorted:
        ic_val = entry.get("IC_published", 0)
        ws.cell(row=row, column=1, value=entry.get("key", ""))
        ws.cell(row=row, column=2, value=entry.get("year", ""))
        ws.cell(row=row, column=3, value="Sim" if entry.get("is_published") else "Não")
        ws.cell(row=row, column=4, value=entry.get("citations_received_from_published", 0))
        ws.cell(row=row, column=5, value=entry.get("n_published_after", 0))
        cell_ic = ws.cell(row=row, column=6, value=round(ic_val, 4) if ic_val else 0)
        cell_ic.number_format = "0.0000"
        ws.cell(row=row, column=7, value=entry.get("publication_type", ""))
        row += 1

    # ===== Column widths =====
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 25

    wb.save(str(OUTPUT_XLSX))
    print(f"Aba Resumo atualizada em {OUTPUT_XLSX}")


if __name__ == "__main__":
    generate_resumo()
