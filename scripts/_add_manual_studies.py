"""Adiciona 10 estudos manuais ao all_papers.xlsx para processamento LLM.

Uso: python scripts/_add_manual_studies.py
"""

import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE_DIR / "data" / "2-papers" / "all_papers.xlsx"

NEW_STUDIES = [
    {
        "base": "manual",
        "url": "http://repositorio.ipea.gov.br/handle/11058/3138",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2014-resende-fno.pdf",
        "id": "",
        "titulo": "Avaliação dos Impactos Regionais do Fundo Constitucional de Financiamento do Norte entre 2004 e 2010",
        "autores": "Resende, Guilherme",
        "ano": 2014,
        "periodico": "Texto para Discussão IPEA n. 1973",
        "doi": "",
        "resumo": "",
        "tipo": "texto para discussão",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese (citacoes_tese_inclusao_manual.md)",
    },
    {
        "base": "manual",
        "url": "https://portalrevistas.ucb.br/index.php/rbee/article/view/6832",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2016-oliveira-lima-arriel.pdf",
        "id": "",
        "titulo": "Fundo Constitucional de Financiamento do Centro-Oeste (FCO) em Goiás: uma Aplicação Econométrica-Espacial",
        "autores": "Oliveira, Guilherme Resende; Lima, Alex Felipe Rodrigues; Arriel, Marcos Fernando",
        "ano": 2016,
        "periodico": "Revista Brasileira de Economia de Empresas, v. 16, n. 1",
        "doi": "",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese. Substitui Oliveira2015 (econpapers-2015).",
    },
    {
        "base": "manual",
        "url": "https://doi.org/10.61673/ren.2017.701",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2017-resende-silva-filho.pdf",
        "id": "",
        "titulo": "Avaliação Econômica do Fundo Constitucional de Financiamento do Nordeste (FNE): uma Análise Espacial por Tipologia da PNDR entre 1999 e 2011",
        "autores": "Resende, Guilherme Mendes; Silva, Danilo Firmino Costa; Silva Filho, Luis Abel",
        "ano": 2017,
        "periodico": "Revista Econômica do Nordeste, v. 48, n. 1, p. 9-29",
        "doi": "10.61673/ren.2017.701",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese.",
    },
    {
        "base": "manual",
        "url": "",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2016-irffi-araujo-bastos.pdf",
        "id": "",
        "titulo": "Efeitos Heterogêneos do Fundo Constitucional de Financiamento do Nordeste na Região do Semiárido",
        "autores": "Irffi, Guilherme; Araújo, Jair Andrade da Silva; Bastos, Fabrício de Souza",
        "ano": 2016,
        "periodico": "Fórum Banco do Nordeste de Desenvolvimento, v. 22",
        "doi": "",
        "resumo": "",
        "tipo": "apresentação em congresso",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese. Publicação interna do BNB.",
    },
    {
        "base": "manual",
        "url": "https://www.bnb.gov.br/s482-dspace/handle/123456789/676",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2018-carneiro.pdf",
        "id": "",
        "titulo": "Determinantes da Eficiência da Aplicação dos Recursos do FNE pelos Municípios Beneficiados",
        "autores": "Carneiro, Diego",
        "ano": 2018,
        "periodico": "Artigos ETENE, Banco do Nordeste do Brasil",
        "doi": "",
        "resumo": "",
        "tipo": "artigo institucional",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese. Série ETENE/BNB.",
    },
    {
        "base": "manual",
        "url": "https://doi.org/10.32586/rcda.v17i1.472",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2019-cambota-viana.pdf",
        "id": "",
        "titulo": "O Impacto do Fundo Constitucional de Financiamento do Nordeste (FNE) no Crescimento dos Municípios: uma Aplicação de Painel Dinâmico",
        "autores": "Cambota, Jacqueline Nogueira; Viana, Leonardo Ferreira Gomes",
        "ano": 2019,
        "periodico": "Revista Controle - Doutrina e Artigos, v. 17, n. 1, p. 20-46",
        "doi": "10.32586/rcda.v17i1.472",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese. Periódico do TCE-CE.",
    },
    {
        "base": "manual",
        "url": "https://doi.org/10.22456/2176-5456.92093",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2022-goncalves-braga-gurgel.pdf",
        "id": "",
        "titulo": "Avaliação dos Impactos do Fundo Constitucional de Financiamento do Nordeste (FNE): uma Abordagem de Equilíbrio Geral",
        "autores": "Gonçalves, Marcelo Ferreira; Braga, Marcelo José; Gurgel, Angelo Costa",
        "ano": 2022,
        "periodico": "Análise Econômica, v. 40, n. 81",
        "doi": "10.22456/2176-5456.92093",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese.",
    },
    {
        "base": "manual",
        "url": "https://doi.org/10.61673/ren.2020.1106",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2020-rieger-lima-rodrigues.pdf",
        "id": "",
        "titulo": "O Efeito do FNE no Crescimento do Emprego Formal da Região Nordeste",
        "autores": "Rieger, Roberto Arruda; Lima, Regina Marta Nepomuceno; Rodrigues, Clauber Teixeira",
        "ano": 2020,
        "periodico": "Revista Econômica do Nordeste, v. 51, n. 2, p. 155-168",
        "doi": "10.61673/ren.2020.1106",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese.",
    },
    {
        "base": "manual",
        "url": "https://www.ipea.gov.br/ppp/index.php/PPP/article/view/945",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2020-daniel-braga.pdf",
        "id": "",
        "titulo": "Impactos do Fundo Constitucional de Financiamento do Norte: evidências do estimador de diferenças em diferenças",
        "autores": "Daniel, Lindomar Pegorini; Braga, Marcelo José",
        "ano": 2020,
        "periodico": "Planejamento e Políticas Públicas (PPP/IPEA), n. 55, p. 97-146",
        "doi": "10.38116/ppp55art4",
        "resumo": "",
        "tipo": "artigo publicado",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese.",
    },
    {
        "base": "manual",
        "url": "https://www.usp.br/nereus/wp-content/uploads/TD_NEREUS_03_2023.pdf",
        "baixado": "Sim",
        "arquivo_pdf": "manual-2023-silva-azzoni-chagas.pdf",
        "id": "",
        "titulo": "Impactos do Financiamento Público sobre a Economia dos Municípios Beneficiados",
        "autores": "da Silva Filho, Luís Abel; Azzoni, Carlos Roberto; Chagas, André Luís Squarize",
        "ano": 2023,
        "periodico": "TD NEREUS n. 03-2023, USP",
        "doi": "",
        "resumo": "",
        "tipo": "texto para discussão",
        "palavras_chave": "",
        "obs": "Inclusão manual via tese. Working paper NEREUS/USP.",
    },
]

COLS = [
    "base", "url", "baixado", "arquivo_pdf", "id", "titulo",
    "autores", "ano", "periodico", "doi", "resumo", "tipo",
    "palavras_chave", "obs",
]


def main() -> None:
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb["Registros"]

    before = ws.max_row - 1
    for study in NEW_STUDIES:
        row_data = [study[c] for c in COLS]
        ws.append(row_data)

    wb.save(str(XLSX_PATH))
    after = ws.max_row - 1
    wb.close()

    print(f"Adicionados {after - before} estudos a all_papers.xlsx")
    print(f"  Antes: {before} rows → Depois: {after} rows")


if __name__ == "__main__":
    main()
