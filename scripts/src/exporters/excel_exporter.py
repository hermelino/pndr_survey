"""Exportação de resultados para Excel (openpyxl)."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exporters.common import flatten_all
from src.models import PaperRecord

logger = logging.getLogger("pndr_survey")


def export_excel(
    papers: List[PaperRecord],
    output_path: str | Path,
) -> Path:
    """Exporta PaperRecords para arquivo Excel.

    Cria duas sheets:
      - "Resultados": uma linha por paper com todos os campos
      - "Resumo": contagens por instrumento PNDR e método econométrico

    Args:
        papers: Lista de PaperRecords.
        output_path: Caminho do arquivo .xlsx de saída.

    Returns:
        Path do arquivo criado.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = flatten_all(papers)
    if not rows:
        logger.warning("Nenhum registro para exportar.")
        return output_path

    wb = openpyxl.Workbook()

    # --- Sheet: Resultados ---
    ws = wb.active
    ws.title = "Resultados"

    headers = list(rows[0].keys())
    _write_header(ws, headers)

    for i, row in enumerate(rows, start=2):
        for j, key in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws, headers, rows)
    ws.freeze_panes = "C2"

    # --- Sheet: Resumo ---
    ws_sum = wb.create_sheet("Resumo")
    _write_summary(ws_sum, papers)

    wb.save(output_path)
    logger.info("Excel exportado: %s (%d registros)", output_path, len(rows))
    return output_path


# --- Helpers ---

_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def _write_header(ws, headers: list) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, headers: list, rows: list) -> None:
    for j, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row in rows[:50]:  # amostrar primeiras 50 linhas
            val = str(row.get(h, ""))
            max_len = max(max_len, min(len(val), 60))
        col_letter = get_column_letter(j)
        ws.column_dimensions[col_letter].width = max_len + 2


def _write_summary(ws, papers: List[PaperRecord]) -> None:
    ws.cell(row=1, column=1, value="Resumo da Revisão Sistemática").font = Font(bold=True, size=12)

    row = 3
    ws.cell(row=row, column=1, value="Total de registros:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(papers))

    empirical = [p for p in papers if p.is_empirical]
    row += 1
    ws.cell(row=row, column=1, value="Estudos empíricos (passaram triagem):").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(empirical))

    # Contagem por instrumento
    row += 2
    ws.cell(row=row, column=1, value="Por instrumento PNDR").font = Font(bold=True)
    _write_header(ws, ["Instrumento", "Contagem"])
    # Ajustar posição do header
    ws.cell(row=row + 1, column=1).value = "Instrumento"
    ws.cell(row=row + 1, column=1).font = Font(bold=True)
    ws.cell(row=row + 1, column=2).value = "Contagem"
    ws.cell(row=row + 1, column=2).font = Font(bold=True)

    inst_counts: dict = {}
    for p in empirical:
        inst = p.pndr_instrument or "Não classificado"
        inst_counts[inst] = inst_counts.get(inst, 0) + 1

    for i, (inst, count) in enumerate(sorted(inst_counts.items()), start=row + 2):
        ws.cell(row=i, column=1, value=inst)
        ws.cell(row=i, column=2, value=count)

    # Contagem por método
    row = row + len(inst_counts) + 4
    ws.cell(row=row, column=1, value="Por método econométrico").font = Font(bold=True)
    ws.cell(row=row + 1, column=1).value = "Método"
    ws.cell(row=row + 1, column=1).font = Font(bold=True)
    ws.cell(row=row + 1, column=2).value = "Contagem"
    ws.cell(row=row + 1, column=2).font = Font(bold=True)

    method_counts: dict = {}
    for p in empirical:
        if p.stage_2 and isinstance(p.stage_2, dict):
            method = p.stage_2.get("metodo_econometrico", "Não informado")
        else:
            method = "Não analisado"
        method_counts[method] = method_counts.get(method, 0) + 1

    for i, (method, count) in enumerate(sorted(method_counts.items()), start=row + 2):
        ws.cell(row=i, column=1, value=method)
        ws.cell(row=i, column=2, value=count)
