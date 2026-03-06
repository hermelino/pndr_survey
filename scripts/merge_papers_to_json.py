"""Mescla all_papers_llm_classif_final.xlsx com registros das bases de dados.

Combina:
  1. all_papers_llm_classif_final.xlsx (triagem + análise LLM S1/S2/S3)
  2. all_papers.xlsx (URL, Resumo, Tipo, Palavras-chave, ID)
  3. bib_records.json (metadados completos das bases: abstract, keywords, volume, etc.)

Para campos em duplicidade, prioridade: bib_records > all_papers > llm_classification.
Salva resultado como JSON em data/2-papers/2-2-papers.json.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz
from unidecode import unidecode

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

LLM_XLSX = DATA_DIR / "2-papers" / "all_papers_llm_classif_final.xlsx"
ALL_PAPERS_XLSX = DATA_DIR / "2-papers" / "all_papers.xlsx"
BIB_RECORDS_JSON = DATA_DIR / "1-records" / "processed" / "bib_records.json"
LLM_CHECKPOINT = DATA_DIR / "2-papers" / "_llm_checkpoint.json"
OUTPUT_JSON = DATA_DIR / "2-papers" / "2-2-papers.json"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
STOPWORDS = {
    "the", "a", "an", "of", "in", "on", "at", "to", "for", "and", "or",
    "with", "by", "from", "as", "is", "was", "are", "were", "been",
    "o", "os", "um", "uma", "uns", "umas", "de", "do", "da",
    "dos", "das", "em", "no", "na", "nos", "nas", "por", "para", "com",
    "que", "se", "ao", "aos",
    "el", "la", "los", "las", "un", "una", "unos", "unas", "del", "en", "con",
}


def normalize_title(title: str) -> str:
    text = unidecode(title).lower()
    text = re.sub(r"[^\w\s]", " ", text)
    words = [w for w in text.split() if w not in STOPWORDS]
    return " ".join(words)


def clean_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def clean_year(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Handle "2024.0" from Excel
    s = s.split(".")[0]
    return s


# ---------------------------------------------------------------------------
# Load data sources
# ---------------------------------------------------------------------------
def load_llm_xlsx() -> dict:
    """Load LLM classification xlsx, keyed by Arquivo PDF."""
    wb = openpyxl.load_workbook(str(LLM_XLSX), read_only=True)
    # Use first sheet (data), not the active sheet (may be a summary tab)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    papers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                record[headers[j]] = val
        pdf = clean_str(record.get("Arquivo PDF"))
        if pdf:
            papers[pdf] = record
    wb.close()
    return papers


def load_all_papers_xlsx() -> dict:
    """Load all_papers.xlsx, keyed by Arquivo PDF."""
    wb = openpyxl.load_workbook(str(ALL_PAPERS_XLSX), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    papers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                record[headers[j]] = val
        pdf = clean_str(record.get("Arquivo PDF"))
        if pdf:
            papers[pdf] = record
    wb.close()
    return papers


def load_bib_records() -> list:
    """Load bib_records.json."""
    with open(BIB_RECORDS_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def load_llm_checkpoint() -> dict:
    """Load LLM checkpoint (keyed by PDF filename)."""
    with open(LLM_CHECKPOINT, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Match bib_records to papers by title+year
# ---------------------------------------------------------------------------
def match_bib_to_pdf(bib_records: list, llm_papers: dict) -> dict:
    """Match each bib_record to a PDF filename via fuzzy title + year match.

    Returns dict: pdf_filename -> bib_record.
    """
    # Pre-compute normalized titles for papers
    paper_norms = {}
    for pdf, rec in llm_papers.items():
        titulo = clean_str(rec.get("Titulo") or rec.get("S1_titulo") or "")
        ano = clean_year(rec.get("Ano"))
        base = clean_str(rec.get("Base")).lower()
        paper_norms[pdf] = {
            "titulo_norm": normalize_title(titulo) if titulo else "",
            "ano": ano,
            "base": base,
        }

    # Pre-compute for bib_records
    bib_norms = []
    for i, br in enumerate(bib_records):
        bib_norms.append({
            "titulo_norm": normalize_title(br.get("title", "")),
            "ano": str(br.get("year", "")),
            "source_db": (br.get("source_db") or "").lower(),
        })

    matched = {}
    used_bib = set()

    for pdf, pn in paper_norms.items():
        if not pn["titulo_norm"]:
            continue

        best_score = 0
        best_idx = None

        for i, bn in enumerate(bib_norms):
            if i in used_bib:
                continue
            if not bn["titulo_norm"]:
                continue

            # Year filter
            if pn["ano"] and bn["ano"] and pn["ano"] != bn["ano"]:
                continue

            # Source DB filter (soft: prefer same source)
            score = fuzz.token_sort_ratio(pn["titulo_norm"], bn["titulo_norm"])

            # Bonus for same source_db
            if pn["base"] and bn["source_db"] and pn["base"] == bn["source_db"]:
                score += 2  # small bonus

            if score > best_score and score >= 80:
                best_score = score
                best_idx = i

        if best_idx is not None:
            matched[pdf] = bib_records[best_idx]
            used_bib.add(best_idx)

    return matched


# ---------------------------------------------------------------------------
# Merge
# ---------------------------------------------------------------------------
def merge_papers(llm_papers: dict, all_papers: dict, bib_matched: dict, checkpoint: dict) -> list:
    """Merge all sources into enriched paper records."""
    result = []

    for pdf, llm in llm_papers.items():
        ap = all_papers.get(pdf, {})
        bib = bib_matched.get(pdf)
        chk = checkpoint.get(pdf, {})

        paper = {}

        # --- Identidade ---
        paper["arquivo_pdf"] = pdf
        paper["num"] = llm.get("#")
        paper["base"] = clean_str(llm.get("Base"))

        # --- Metadados bibliográficos (prioridade: bib > ap > llm) ---
        # Titulo
        paper["titulo"] = (
            clean_str(bib.get("title")) if bib and bib.get("title")
            else clean_str(ap.get("Titulo")) if ap.get("Titulo")
            else clean_str(llm.get("Titulo"))
        )

        # Autores
        paper["autores"] = (
            clean_str(bib.get("authors")) if bib and bib.get("authors")
            else clean_str(ap.get("Autores")) if ap.get("Autores")
            else clean_str(llm.get("Autores"))
        )

        # Ano
        paper["ano"] = (
            str(bib["year"]) if bib and bib.get("year")
            else clean_year(ap.get("Ano")) if ap.get("Ano")
            else clean_year(llm.get("Ano"))
        )

        # Periodico
        paper["periodico"] = (
            clean_str(bib.get("journal")) if bib and bib.get("journal")
            else clean_str(ap.get("Periodico")) if ap.get("Periodico")
            else clean_str(llm.get("Periodico"))
        )

        # DOI
        paper["doi"] = (
            clean_str(bib.get("doi")) if bib and bib.get("doi")
            else clean_str(ap.get("DOI")) if ap.get("DOI")
            else clean_str(llm.get("DOI"))
        )

        # --- Campos extras dos registros (bib_records / all_papers) ---
        paper["url"] = (
            clean_str(bib.get("url")) if bib and bib.get("url")
            else clean_str(ap.get("URL"))
        )
        paper["id_registro"] = clean_str(ap.get("ID"))

        paper["resumo"] = (
            clean_str(bib.get("abstract")) if bib and bib.get("abstract")
            else clean_str(ap.get("Resumo"))
        )

        paper["palavras_chave"] = (
            clean_str(bib.get("keywords")) if bib and bib.get("keywords")
            else clean_str(ap.get("Palavras-chave"))
        )

        paper["tipo_publicacao"] = (
            clean_str(bib.get("publication_type")) if bib and bib.get("publication_type")
            else clean_str(ap.get("Tipo"))
        )

        # Campos exclusivos de bib_records
        if bib:
            paper["volume"] = clean_str(bib.get("volume"))
            paper["issue"] = clean_str(bib.get("issue"))
            paper["pages"] = clean_str(bib.get("pages"))
            paper["pdf_url"] = clean_str(bib.get("pdf_url"))
            paper["idioma"] = clean_str(bib.get("language"))
            paper["instrumentos_detectados"] = clean_str(bib.get("matched_instruments"))
            paper["source_db"] = clean_str(bib.get("source_db"))
            paper["is_duplicate"] = bib.get("is_duplicate", False)
            paper["duplicate_of"] = clean_str(bib.get("duplicate_of"))
        else:
            paper["volume"] = ""
            paper["issue"] = ""
            paper["pages"] = ""
            paper["pdf_url"] = ""
            paper["idioma"] = ""
            paper["instrumentos_detectados"] = ""
            paper["source_db"] = paper["base"].lower()
            paper["is_duplicate"] = False
            paper["duplicate_of"] = ""

        paper["obs"] = clean_str(ap.get("Obs"))

        # --- Triagem ---
        paper["triagem"] = clean_str(llm.get("Triagem"))
        paper["motivo_exclusao"] = clean_str(llm.get("Motivo Exclusão"))

        # --- LLM Stage 1 ---
        paper["s1"] = {
            "is_scientific_study": clean_str(llm.get("S1_is_scientific_study")),
            "tipo_trabalho": clean_str(llm.get("S1_tipo_trabalho")),
            "revista": clean_str(llm.get("S1_revista")),
            "titulo": clean_str(llm.get("S1_titulo")),
            "autores": clean_str(llm.get("S1_autores")),
            "ano_estudo": clean_str(llm.get("S1_ano_estudo")),
            "instrumentos_pndr": clean_str(llm.get("S1_instrumentos_pndr")),
            "metodologia": clean_str(llm.get("S1_metodologia")),
            "uso_econometria": clean_str(llm.get("S1_uso_econometria")),
            "questao_chave": clean_str(llm.get("S1_questao_chave")),
        }

        # --- LLM Stage 2 ---
        paper["s2"] = {
            "metodo_econometrico": clean_str(llm.get("S2_metodo_econometrico")),
            "detalhes_metodo": clean_str(llm.get("S2_detalhes_metodo")),
            "tipo_dados": clean_str(llm.get("S2_tipo_dados")),
            "var_dependente": clean_str(llm.get("S2_var_dependente")),
            "var_controle": clean_str(llm.get("S2_var_controle")),
            "setor_economico": clean_str(llm.get("S2_setor_economico")),
            "area_geografica": clean_str(llm.get("S2_area_geografica")),
            "unidade_tempo": clean_str(llm.get("S2_unidade_tempo")),
            "unidade_espacial": clean_str(llm.get("S2_unidade_espacial")),
            "periodo_amostra": clean_str(llm.get("S2_periodo_amostra")),
        }

        # --- LLM Stage 3 ---
        paper["s3"] = {
            "efeito_parcial": clean_str(llm.get("S3_efeito_parcial")),
            "significancia": clean_str(llm.get("S3_significancia")),
            "direcao_efeito": clean_str(llm.get("S3_direcao_efeito")),
            "outros_resultados": clean_str(llm.get("S3_outros_resultados")),
            "implicacoes": clean_str(llm.get("S3_implicacoes")),
            "limitacoes": clean_str(llm.get("S3_limitacoes")),
            "sugestoes": clean_str(llm.get("S3_sugestoes")),
        }

        # --- Erros de processamento ---
        paper["erros"] = clean_str(llm.get("Erros"))

        # --- Text length from checkpoint ---
        paper["text_length"] = chk.get("text_length", 0)

        result.append(paper)

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Carregando dados...")

    print("  1. all_papers_llm_classif_final.xlsx")
    llm_papers = load_llm_xlsx()
    print(f"     {len(llm_papers)} papers")

    print("  2. all_papers.xlsx")
    all_papers = load_all_papers_xlsx()
    print(f"     {len(all_papers)} papers")

    print("  3. bib_records.json")
    bib_records = load_bib_records()
    print(f"     {len(bib_records)} registros")

    print("  4. _llm_checkpoint.json")
    checkpoint = load_llm_checkpoint()
    print(f"     {len(checkpoint)} entries")

    print("\nMatching bib_records -> papers por título+ano...")
    bib_matched = match_bib_to_pdf(bib_records, llm_papers)
    print(f"  {len(bib_matched)}/{len(llm_papers)} papers matched com bib_records")

    # Report unmatched
    unmatched = set(llm_papers.keys()) - set(bib_matched.keys())
    if unmatched:
        print(f"  Sem match em bib_records ({len(unmatched)}):")
        for pdf in sorted(unmatched):
            print(f"    - {pdf}")

    print("\nMesclando dados...")
    result = merge_papers(llm_papers, all_papers, bib_matched, checkpoint)

    # Stats
    from collections import Counter
    triagem = Counter(p["triagem"] for p in result)
    with_abstract = sum(1 for p in result if p["resumo"])
    with_keywords = sum(1 for p in result if p["palavras_chave"])
    with_bib = sum(1 for p in result if p.get("volume") or p.get("issue") or p.get("pages"))

    print(f"\nResultado: {len(result)} papers")
    print(f"  Triagem: {dict(triagem)}")
    print(f"  Com resumo: {with_abstract}")
    print(f"  Com palavras-chave: {with_keywords}")
    print(f"  Com volume/issue/pages: {with_bib}")

    print(f"\nSalvando em {OUTPUT_JSON}...")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print("Concluído!")


if __name__ == "__main__":
    main()
