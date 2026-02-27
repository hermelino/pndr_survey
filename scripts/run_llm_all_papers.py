"""Análise LLM completa de todos os papers em all_papers.xlsx.

Pipeline:
  1. Lê all_papers.xlsx (118 registros, todos com PDF)
  2. Extrai texto de cada PDF
  3. Stage 1 (triagem): 10 perguntas → filtra por is_scientific_study + instrumentos + econometria
  4. Stage 2 (metodologia): 10 perguntas nos que passam
  5. Stage 3 (resultados): 7 perguntas nos que passam
  6. Salva tudo em all_papers_llm_classification.xlsx

Inclui checkpoint JSON para retomar em caso de falha.
"""

import json
import logging
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Adicionar scripts/ ao sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.config import load_config
from src.models import BibRecord
from src.extractors.pdf_extractor import PdfExtractor
from src.analyzers.gemini_analyzer import GeminiAnalyzer, filter_screening
from src.analyzers.base import load_questionnaire

# --- Caminhos ---
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data" / "2-papers"
PDF_DIR = DATA_DIR / "2-2-papers-pdfs"
INPUT_XLSX = DATA_DIR / "all_papers.xlsx"
OUTPUT_XLSX = DATA_DIR / "all_papers_llm_classification.xlsx"
CHECKPOINT_FILE = DATA_DIR / "_llm_checkpoint.json"

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(DATA_DIR / "_llm_run.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("pndr_survey")


def load_checkpoint():
    """Carrega checkpoint de execução anterior."""
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_checkpoint(data):
    """Salva checkpoint incrementalmente."""
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def read_papers_from_xlsx():
    """Lê registros do all_papers.xlsx."""
    wb = openpyxl.load_workbook(INPUT_XLSX, read_only=True)
    ws = wb["Registros"]

    papers = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        papers.append({
            "base": row[0] or "",
            "url": row[1] or "",
            "baixado": row[2] or "",
            "arquivo_pdf": row[3] or "",
            "id": row[4] or "",
            "titulo": row[5] or "",
            "autores": row[6] or "",
            "ano": row[7],
            "periodico": row[8] or "",
            "doi": row[9] or "",
            "resumo": row[10] or "",
            "tipo": row[11] or "",
            "palavras_chave": row[12] or "",
            "obs": row[13] or "",
        })
    wb.close()
    return papers


def run_analysis():
    """Executa pipeline completo de análise LLM."""
    # --- Config e componentes ---
    config = load_config(SCRIPT_DIR / "config.yaml")
    extractor = PdfExtractor(max_chars=config.llm.max_tokens_input)
    # gemini-2.5-flash-lite: 1000 RPD free tier (vs 20 RPD do 2.5-flash)
    model_name = "gemini-2.5-flash-lite"
    analyzer = GeminiAnalyzer(
        api_key=config.llm.api_key,
        model=model_name,
        temperature=config.llm.temperature,
        max_tokens_input=config.llm.max_tokens_input,
        rate_limit_seconds=max(config.llm.rate_limit_seconds, 6.0),
        max_retries=5,
    )

    q1 = load_questionnaire(SCRIPT_DIR / "questionnaires" / "stage_1_screening.json")
    q2 = load_questionnaire(SCRIPT_DIR / "questionnaires" / "stage_2_methods.json")
    q3 = load_questionnaire(SCRIPT_DIR / "questionnaires" / "stage_3_results.json")

    q1_ids = [q["id"] for q in q1["questions"]]
    q2_ids = [q["id"] for q in q2["questions"]]
    q3_ids = [q["id"] for q in q3["questions"]]

    # --- Ler papers ---
    papers = read_papers_from_xlsx()
    logger.info("Total de papers: %d", len(papers))

    # --- Checkpoint ---
    checkpoint = load_checkpoint()

    # Limpar entradas com erro para reprocessar
    error_keys = [k for k, v in checkpoint.items()
                  if v.get("error") or v.get("error_stage_1") or v.get("error_stage_2") or v.get("error_stage_3")]
    for k in error_keys:
        del checkpoint[k]
    if error_keys:
        save_checkpoint(checkpoint)
        logger.info("Removidas %d entradas com erro para reprocessamento", len(error_keys))

    logger.info("Checkpoint: %d papers já processados com sucesso", len(checkpoint))

    # --- Processar cada paper ---
    start_time = time.time()
    total = len(papers)
    stage1_ok = 0
    stage1_fail = 0
    passed_screening = 0

    for i, paper_info in enumerate(papers):
        pdf_name = paper_info["arquivo_pdf"]
        key = pdf_name  # Chave única do checkpoint

        # Pular se já processado
        if key in checkpoint:
            s1 = checkpoint[key].get("stage_1")
            if s1:
                stage1_ok += 1
                if checkpoint[key].get("passed_screening"):
                    passed_screening += 1
            continue

        pdf_path = PDF_DIR / pdf_name
        if not pdf_path.exists():
            logger.warning("[%d/%d] PDF não encontrado: %s", i + 1, total, pdf_name)
            checkpoint[key] = {"error": f"PDF não encontrado: {pdf_name}"}
            save_checkpoint(checkpoint)
            continue

        logger.info("[%d/%d] Processando: %s", i + 1, total, pdf_name)

        # --- Extrair texto ---
        bib = BibRecord(
            source_db=paper_info["base"],
            source_id=paper_info["id"] or pdf_name,
            title=paper_info["titulo"],
            authors=paper_info["autores"].split(";") if paper_info["autores"] else [],
            year=paper_info["ano"],
            doi=paper_info["doi"] or None,
        )
        paper_rec = extractor.extract(pdf_path, bib)

        if not getattr(paper_rec, "_extracted_text", None):
            logger.warning("  Sem texto extraído: %s", pdf_name)
            checkpoint[key] = {"error": "Sem texto extraído"}
            save_checkpoint(checkpoint)
            continue

        result = {"text_length": paper_rec.text_length}

        # --- Stage 1: Triagem ---
        try:
            s1 = analyzer.analyze(paper_rec, q1)
            paper_rec.stage_1 = s1
            result["stage_1"] = s1
            stage1_ok += 1
            logger.info("  Stage 1 OK: is_study=%s, instrumentos=%s, econometria=%s",
                        s1.get("is_scientific_study", "?")[:20],
                        s1.get("instrumentos_pndr", "?")[:30],
                        s1.get("uso_econometria", "?")[:10])
        except Exception as e:
            logger.error("  Stage 1 FALHOU: %s", e)
            result["error_stage_1"] = str(e)
            stage1_fail += 1
            checkpoint[key] = result
            save_checkpoint(checkpoint)
            continue

        # --- Filtro de triagem ---
        passed_list, _ = filter_screening([paper_rec])
        did_pass = len(passed_list) > 0
        result["passed_screening"] = did_pass

        if not did_pass:
            reason = paper_rec.bib.exclusion_reason or "Não passou na triagem"
            result["exclusion_reason"] = reason
            logger.info("  Triagem: REJEITADO (%s)", reason)
            checkpoint[key] = result
            save_checkpoint(checkpoint)
            continue

        passed_screening += 1
        logger.info("  Triagem: APROVADO → Stages 2 e 3")

        # --- Stage 2: Metodologia ---
        try:
            s2 = analyzer.analyze(paper_rec, q2)
            paper_rec.stage_2 = s2
            result["stage_2"] = s2
            logger.info("  Stage 2 OK: metodo=%s", s2.get("metodo_econometrico", "?")[:40])
        except Exception as e:
            logger.error("  Stage 2 FALHOU: %s", e)
            result["error_stage_2"] = str(e)

        # --- Stage 3: Resultados ---
        try:
            s3 = analyzer.analyze(paper_rec, q3)
            paper_rec.stage_3 = s3
            result["stage_3"] = s3
            logger.info("  Stage 3 OK: efeito=%s, direcao=%s",
                        s3.get("efeito_parcial", "?")[:30],
                        s3.get("direcao_efeito", "?")[:20])
        except Exception as e:
            logger.error("  Stage 3 FALHOU: %s", e)
            result["error_stage_3"] = str(e)

        checkpoint[key] = result
        save_checkpoint(checkpoint)

        # Progresso
        elapsed = time.time() - start_time
        rate = (i + 1) / elapsed if elapsed > 0 else 0
        remaining = (total - i - 1) / rate if rate > 0 else 0
        logger.info("  Progresso: %d/%d | Stage1 ok=%d fail=%d | Aprovados=%d | ETA: %.0fm",
                     i + 1, total, stage1_ok, stage1_fail, passed_screening, remaining / 60)

    # --- Gerar XLSX ---
    logger.info("=" * 60)
    logger.info("Análise concluída. Gerando %s", OUTPUT_XLSX.name)
    write_output_xlsx(papers, checkpoint, q1_ids, q2_ids, q3_ids)

    logger.info("=" * 60)
    logger.info("RESUMO FINAL:")
    logger.info("  Total papers: %d", total)
    logger.info("  Stage 1 analisados: %d", stage1_ok)
    logger.info("  Aprovados (Stage 2+3): %d", passed_screening)
    logger.info("  Arquivo: %s", OUTPUT_XLSX)


def write_output_xlsx(papers, checkpoint, q1_ids, q2_ids, q3_ids):
    """Gera all_papers_llm_classification.xlsx com todos os resultados."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classificação LLM"

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill_meta = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_s1 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_fill_filter = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    header_fill_s2 = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_fill_s3 = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    green_font = Font(color="006400")
    red_font = Font(color="8B0000")

    # --- Headers ---
    meta_headers = ["#", "Base", "Arquivo PDF", "Titulo", "Autores", "Ano", "Periodico", "DOI"]
    s1_headers = [f"S1_{qid}" for qid in q1_ids]
    filter_headers = ["Triagem", "Motivo Exclusão"]
    s2_headers = [f"S2_{qid}" for qid in q2_ids]
    s3_headers = [f"S3_{qid}" for qid in q3_ids]
    error_headers = ["Erros"]

    all_headers = meta_headers + s1_headers + filter_headers + s2_headers + s3_headers + error_headers

    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = wrap_align

        # Colorir por seção
        if col <= len(meta_headers):
            cell.fill = header_fill_meta
        elif col <= len(meta_headers) + len(s1_headers):
            cell.fill = header_fill_s1
        elif col <= len(meta_headers) + len(s1_headers) + len(filter_headers):
            cell.fill = header_fill_filter
        elif col <= len(meta_headers) + len(s1_headers) + len(filter_headers) + len(s2_headers):
            cell.fill = header_fill_s2
        elif col <= len(meta_headers) + len(s1_headers) + len(filter_headers) + len(s2_headers) + len(s3_headers):
            cell.fill = header_fill_s3
        else:
            cell.fill = header_fill_meta

    # --- Data rows ---
    for i, paper_info in enumerate(papers):
        row = i + 2
        pdf_name = paper_info["arquivo_pdf"]
        cp = checkpoint.get(pdf_name, {})

        col = 1

        # Meta
        for val in [
            i + 1,
            paper_info["base"],
            pdf_name,
            paper_info["titulo"],
            paper_info["autores"],
            paper_info["ano"],
            paper_info["periodico"],
            paper_info["doi"],
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = wrap_align
            col += 1

        # Stage 1
        s1 = cp.get("stage_1", {})
        for qid in q1_ids:
            ws.cell(row=row, column=col, value=s1.get(qid, "")).alignment = wrap_align
            col += 1

        # Filtro triagem
        passed = cp.get("passed_screening")
        if passed is True:
            cell = ws.cell(row=row, column=col, value="APROVADO")
            cell.font = green_font
        elif passed is False:
            cell = ws.cell(row=row, column=col, value="REJEITADO")
            cell.font = red_font
        elif cp.get("error"):
            ws.cell(row=row, column=col, value="ERRO")
        else:
            ws.cell(row=row, column=col, value="")
        col += 1

        ws.cell(row=row, column=col, value=cp.get("exclusion_reason", "")).alignment = wrap_align
        col += 1

        # Stage 2
        s2 = cp.get("stage_2", {})
        for qid in q2_ids:
            ws.cell(row=row, column=col, value=s2.get(qid, "")).alignment = wrap_align
            col += 1

        # Stage 3
        s3 = cp.get("stage_3", {})
        for qid in q3_ids:
            ws.cell(row=row, column=col, value=s3.get(qid, "")).alignment = wrap_align
            col += 1

        # Erros
        errors = []
        if cp.get("error"):
            errors.append(cp["error"])
        if cp.get("error_stage_1"):
            errors.append(f"S1: {cp['error_stage_1']}")
        if cp.get("error_stage_2"):
            errors.append(f"S2: {cp['error_stage_2']}")
        if cp.get("error_stage_3"):
            errors.append(f"S3: {cp['error_stage_3']}")
        ws.cell(row=row, column=col, value="; ".join(errors)).alignment = wrap_align

    # --- Larguras de coluna ---
    ws.column_dimensions["A"].width = 4    # #
    ws.column_dimensions["B"].width = 12   # Base
    ws.column_dimensions["C"].width = 40   # Arquivo PDF
    ws.column_dimensions["D"].width = 55   # Titulo
    ws.column_dimensions["E"].width = 40   # Autores
    ws.column_dimensions["F"].width = 6    # Ano
    ws.column_dimensions["G"].width = 25   # Periodico
    ws.column_dimensions["H"].width = 25   # DOI

    # Freeze header and meta columns
    ws.freeze_panes = "I2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet Resumo ---
    ws_resumo = wb.create_sheet("Resumo")
    total = len(papers)
    analyzed = sum(1 for p in papers if checkpoint.get(p["arquivo_pdf"], {}).get("stage_1"))
    approved = sum(1 for p in papers if checkpoint.get(p["arquivo_pdf"], {}).get("passed_screening"))
    rejected = analyzed - approved
    errors = sum(1 for p in papers if checkpoint.get(p["arquivo_pdf"], {}).get("error"))

    summary = [
        ("Métrica", "Valor"),
        ("Total de papers", total),
        ("Analisados (Stage 1)", analyzed),
        ("Aprovados (Stages 2+3)", approved),
        ("Rejeitados na triagem", rejected),
        ("Erros de processamento", errors),
        ("", ""),
        ("Motivo de Exclusão", "Qtd"),
    ]

    # Contar motivos de exclusão
    exclusion_counts = {}
    for p in papers:
        cp = checkpoint.get(p["arquivo_pdf"], {})
        reason = cp.get("exclusion_reason", "")
        if reason:
            exclusion_counts[reason] = exclusion_counts.get(reason, 0) + 1

    for reason, count in sorted(exclusion_counts.items(), key=lambda x: -x[1]):
        summary.append((reason, count))

    for r, (label, val) in enumerate(summary, 1):
        ws_resumo.cell(row=r, column=1, value=label).font = Font(bold=(r == 1 or r == 8))
        ws_resumo.cell(row=r, column=2, value=val).font = Font(bold=(r == 1 or r == 8))

    ws_resumo.column_dimensions["A"].width = 35
    ws_resumo.column_dimensions["B"].width = 12

    wb.save(OUTPUT_XLSX)
    logger.info("Arquivo salvo: %s", OUTPUT_XLSX)


if __name__ == "__main__":
    run_analysis()
