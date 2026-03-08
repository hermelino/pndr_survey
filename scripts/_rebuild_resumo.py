"""Generate resumo_classificacao.xlsx from all_papers_llm_classif_final.xlsx data sheet.

Reads the 'Classificacao LLM' sheet (source of truth) and produces a standalone
summary Excel file with statistics. This avoids modifying the manual classification
file programmatically.
"""

from pathlib import Path
from collections import Counter

import openpyxl
import rispy
from openpyxl.styles import Font, Alignment
from unidecode import unidecode

INPUT_PATH = Path(__file__).parent.parent / "data" / "2-papers" / "all_papers_llm_classif_final.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent / "data" / "2-papers" / "resumo_classificacao.xlsx"

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")

# ---------------------------------------------------------------------------
# Motivos de exclusao
# ---------------------------------------------------------------------------
MOTIVO_KEYWORDS: list[tuple[list[str], str]] = [
    (["instrumento", "pndr"], "sem instrumentos PNDR"),
    (["metodo econom"], "sem metodo econometrico"),
    (["estudo cient", "trabalho cient"], "nao e estudo cientifico"),
    (["duplicata"], "duplicata de versao publicada"),
    (["anterior"], "anterior a 2005"),
    (["escopo", "variav"], "variaveis de resultado fora do escopo"),
]

MOTIVO_ORDER = [m for _, m in MOTIVO_KEYWORDS]


def normalize_motivo(raw: str) -> str:
    """Map raw exclusion reason to canonical form via keyword matching."""
    key = unidecode(raw.strip().lower())
    for keywords, canonical in MOTIVO_KEYWORDS:
        for kw in keywords:
            if kw in key:
                return canonical
    return raw.strip()


# ---------------------------------------------------------------------------
# Instrumentos PNDR
# ---------------------------------------------------------------------------
INSTRUMENTO_ORDER = ["FNE", "FNO", "FCO", "FDNE", "FDA", "FDCO", "IF -- Sudene", "IF -- Sudam"]

INSTRUMENTO_KEYWORDS: dict[str, list[str]] = {
    "FNE": ["FNE", "Fundo Constitucional de Financiamento do Nordeste"],
    "FNO": ["FNO", "Fundo Constitucional de Financiamento do Norte"],
    "FCO": ["FCO", "Fundo Constitucional de Financiamento do Centro-Oeste"],
    "FDNE": ["FDNE", "Fundo de Desenvolvimento do Nordeste"],
    "FDA": ["FDA", "Fundo de Desenvolvimento da Amaz"],
    "FDCO": ["FDCO", "Fundo de Desenvolvimento do Centro-Oeste"],
    "IF -- Sudene": ["SUDENE", "incentiv"],
    "IF -- Sudam": ["SUDAM"],
}


def parse_instrumentos(text: str | None) -> list[str]:
    """Extract PNDR instruments via keyword search."""
    if not text:
        return []
    s = str(text)
    found = []
    for instr, keywords in INSTRUMENTO_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in s.lower():
                found.append(instr)
                break
    return found


# ---------------------------------------------------------------------------
# Periodos
# ---------------------------------------------------------------------------
PERIODO_BINS = [
    ("2005--2010", 2005, 2010),
    ("2011--2015", 2011, 2015),
    ("2016--2020", 2016, 2020),
    ("2021--2026", 2021, 2026),
]

# ---------------------------------------------------------------------------
# Unidade amostral
# ---------------------------------------------------------------------------
UNIDADE_ORDER = ["Municipio", "Empresa", "Microrregiao", "UF", "Mesorregiao", "Area Minima Comparavel"]

UNIDADE_KEYWORDS: list[tuple[list[str], str]] = [
    (["munic"], "Municipio"),
    (["empresa", "firma"], "Empresa"),
    (["microrr"], "Microrregiao"),
    (["mesorr"], "Mesorregiao"),
    (["area minim", "amc"], "Area Minima Comparavel"),
    (["uf", "estado"], "UF"),
]


def normalize_unidade(raw: str) -> str:
    """Normalize spatial unit to canonical form."""
    key = unidecode(raw.strip().lower())
    for keywords, canonical in UNIDADE_KEYWORDS:
        for kw in keywords:
            if kw in key:
                return canonical
    # Fallback: return cleaned original (strip page refs like "[p. 155]")
    import re
    return re.sub(r"\s*\[.*?\]\s*", "", raw).strip()


# ---------------------------------------------------------------------------
# Metodos econometricos — keyword-based extraction from raw LLM text
# ---------------------------------------------------------------------------
# Order matters: more specific patterns must come BEFORE generic ones.
# Each tuple: (keywords_to_search, canonical_name, MSM_scale)
METODO_RULES: list[tuple[list[str], str, str]] = [
    # Staggered DiD (must come before plain DiD)
    (["escalonado", "staggered", "callaway", "dois estagios", "two-stage"],
     "Diferencas em Diferencas Escalonado", "3"),
    # RDD
    (["descontinua", "discontinuity", "rdd", "grdd"],
     "Regressao Descontinua (RDD)", "4"),
    # GPS (must come before PSM)
    (["generalized propensity", "gps", "escore de propensao generalizado",
      "propensao generalizado", "dose-response"],
     "Generalized Propensity Score (GPS)", "3"),
    # PSM
    (["propensity score matching", "psm", "pareamento por escore"],
     "Propensity Score Matching (PSM)", "3"),
    # Generalized Synthetic Control
    (["synthetic control", "controle sintetico", "sintetico generalizado"],
     "Controle Sintetico Generalizado", "3"),
    # IV
    (["instrumental", " iv "],
     "Variaveis Instrumentais (IV)", "3"),
    # DEA
    (["dea", "envoltoria"],
     "Analise Envoltoria de Dados (DEA)", "n.c."),
    # SFA
    (["fronteira estocastica", "sfa", "stochastic frontier"],
     "Fronteira Estocastica (SFA)", "n.c."),
    # Malmquist
    (["malmquist"],
     "Indice de Malmquist", "n.c."),
    # Threshold
    (["limiar", "threshold"],
     "Modelo de Efeito Limiar (Threshold)", "3"),
    # DiD (plain)
    (["diferencas em diferencas", "diff-in-diff", "did", "differences"],
     "Diferencas em Diferencas (DiD)", "3"),
    # Spatial models: error + panel → unified category
    (["erro espacial", "error espacial", "sdem", "sem ",
      "painel espacial", "spatial durbin", "sdm", "espacial aplicada a dados em painel",
      "modelos espaciais de painel", "espaciais de painel"],
     "Painel Espacial", "3"),
    # AEDE
    (["exploratoria espacial", "aede"],
     "Analise Exploratoria Espacial (AEDE)", "n.c."),
    # CGE
    (["equilibrio geral", "cge", "egc", "computable general"],
     "Equilibrio Geral Computavel (EGC)", "n.c."),
    # Dynamic panel (without GMM keyword)
    (["dinamico", "dynamic"],
     "Painel Dinamico", "3"),
    # Random effects panel
    (["efeitos aleatorios", "random effect"],
     "Painel de Efeitos Aleatorios", "3"),
    # Fixed effects panel (must come after spatial panel)
    (["efeito fixo", "efeitos fixos", "fixed effect"],
     "Painel de Efeitos Fixos", "3"),
    # FD
    (["first-differenc", "primeiras diferenc", " fd "],
     "Primeiras Diferencas (FD)", "3"),
    # OLS / MQO
    (["mqo", "ols", "minimos quadrados"],
     "MQO/OLS", "2"),
    # Quantile regression
    (["quantilica", "quantile"],
     "Regressao Quantilica", "3"),
]

MSM_SCALE: dict[str, str] = {name: msm for _, name, msm in METODO_RULES}


def extract_metodos(raw: str | None) -> list[str]:
    """Extract all canonical method names from a raw LLM description.

    Applies hierarchy rules to avoid double-counting within the same
    method family (e.g., DiD Escalonado suppresses plain DiD).
    """
    if not raw:
        return []
    text = unidecode(str(raw).lower())
    found: list[str] = []
    for keywords, canonical, _ in METODO_RULES:
        for kw in keywords:
            if kw.lower() in text:
                if canonical not in found:
                    found.append(canonical)
                break
    # Hierarchy: specific variant suppresses generic variant
    DID_ESC = "Diferencas em Diferencas Escalonado"
    DID_PLAIN = "Diferencas em Diferencas (DiD)"
    if DID_ESC in found and DID_PLAIN in found:
        found.remove(DID_PLAIN)
    return found


# ---------------------------------------------------------------------------
# Autores — RIS-based with normalization (consistent with generate_latex_tables.py)
# ---------------------------------------------------------------------------
RIS_PATH = Path(__file__).parent.parent / "data" / "2-papers" / "approved_papers.ris"

AUTOR_VARIANTES: dict[str, str] = {
    # Irffi
    'IRFFI, Guilherme': 'Irffi, G.D.',
    'Irffi, G.': 'Irffi, G.D.',
    'Irffi, Guilherme Diniz': 'Irffi, G.D.',
    'IRFFI, Guilherme Diniz': 'Irffi, G.D.',
    # Carneiro
    'CARNEIRO, Diego': 'Carneiro, D.R.F.',
    'Carneiro, D.R.F.': 'Carneiro, D.R.F.',
    'Carneiro, Diego Rafael Fonseca': 'Carneiro, D.R.F.',
    'CARNEIRO, Diego Rafael Fonseca': 'Carneiro, D.R.F.',
    # Resende
    'Mendes Resende, G.': 'Resende, G.M.',
    'Resende, Guilherme': 'Resende, G.M.',
    'Resende, Guilherme Mendes': 'Resende, G.M.',
    'RESENDE, Guilherme Mendes': 'Resende, G.M.',
    # Oliveira G.R.
    'Oliveira, Guilherme Resende': 'Oliveira, G.R.',
    'OLIVEIRA, Guilherme Resende': 'Oliveira, G.R.',
    'Oliveira, G.R.': 'Oliveira, G.R.',
    # Oliveira T.G.
    'OLIVEIRA, Tássia Germano de': 'Oliveira, T.G.',
    # Veloso
    'VELOSO, Pedro': 'Veloso, P.A.S.',
    'Veloso, P.A.S.': 'Veloso, P.A.S.',
    'VELOOSO, Pedro Alexandre Santos': 'Veloso, P.A.S.',
    # Silveira Neto
    'SILVEIRA NETO, Raul da Mota': 'Silveira Neto, R.M.',
    'NETO, Raul da Mota Silveira': 'Silveira Neto, R.M.',
    # Costa
    'COSTA, Edward': 'Costa, E.M.',
    'Costa, E.M.': 'Costa, E.M.',
    'COSTA, Edward Martins': 'Costa, E.M.',
    # Braz
    'BRAZ, Marleton Souza': 'Braz, M.S.',
    'Braz, M.S.': 'Braz, M.S.',
    'BRAZ, Marleton': 'Braz, M.S.',
    # Bastos
    'BASTOS, Felipe': 'Bastos, F.S.',
    'BASTOS, Felipe de Sousa': 'Bastos, F.S.',
    'de Sousa Bastos, F.': 'Bastos, F.S.',
    # Alves
    'ALVES, Denis Fernandes': 'Alves, D.F.',
    # Shirasu
    'SHIRASU, Maitê': 'Shirasu, M.',
    # Soares
    'Ricardo Brito Soares': 'Soares, R.B.',
    'Soares, Ricardo Brito': 'Soares, R.B.',
    'Soares, R.B.': 'Soares, R.B.',
}


def normalizar_autor_ris(autor: str) -> str:
    """Normalize RIS author name to canonical form via variant mapping."""
    return AUTOR_VARIANTES.get(autor.strip(), autor.strip())


def load_autores_from_ris() -> Counter[str]:
    """Load and count authors from approved_papers.ris with normalization."""
    autor_counts: Counter[str] = Counter()
    with open(RIS_PATH, 'r', encoding='utf-8') as f:
        ris_records = list(rispy.load(f))
    for rec in ris_records:
        authors = rec.get('authors', [])
        if isinstance(authors, list):
            for autor in authors:
                if autor:
                    autor_counts[normalizar_autor_ris(autor)] += 1
    return autor_counts


# ---------------------------------------------------------------------------
# Excel writer helpers
# ---------------------------------------------------------------------------
def write_header(ws: openpyxl.worksheet.worksheet.Worksheet,
                 row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD


def write_col_header(ws: openpyxl.worksheet.worksheet.Worksheet,
                     row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD
    cell.alignment = CENTER


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print(f"Reading {INPUT_PATH}...")
    wb_in = openpyxl.load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws_data = wb_in[wb_in.sheetnames[0]]

    # --- Read all data rows ---
    header_row = next(ws_data.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    col_idx: dict[str, int] = {h: i for i, h in enumerate(headers) if h}

    rows: list[dict] = []
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))

    wb_in.close()

    total = len(rows)
    aprovados = [r for r in rows if str(r.get("Triagem", "")).upper() == "APROVADO"]
    rejeitados = [r for r in rows if str(r.get("Triagem", "")).upper() == "REJEITADO"]
    n_aprov = len(aprovados)
    n_rej = len(rejeitados)

    print(f"Total: {total}, Aprovados: {n_aprov}, Rejeitados: {n_rej}")

    # --- Find column for Motivo (may have accent) ---
    motivo_key = None
    for h in headers:
        if h and "motivo" in unidecode(str(h)).lower():
            motivo_key = h
            break

    # --- Motivos de exclusao ---
    motivos: Counter[str] = Counter()
    for r in rejeitados:
        raw = str(r.get(motivo_key, "") or "") if motivo_key else ""
        if raw:
            motivos[normalize_motivo(raw)] += 1

    print(f"Motivos: {dict(motivos)}")
    total_motivos = sum(motivos.values())
    if total_motivos != n_rej:
        print(f"  AVISO: soma motivos ({total_motivos}) != rejeitados ({n_rej})")

    # --- Aprovados por periodo ---
    periodo_counts: dict[str, int] = {}
    for label, y_min, y_max in PERIODO_BINS:
        count = sum(1 for r in aprovados if y_min <= int(r.get("Ano", 0) or 0) <= y_max)
        periodo_counts[label] = count

    print(f"Periodos: {periodo_counts}, soma={sum(periodo_counts.values())}")

    # --- Instrumentos PNDR ---
    instr_counts: Counter[str] = Counter()
    for r in aprovados:
        for instr in parse_instrumentos(r.get("S1_instrumentos_pndr")):
            instr_counts[instr] += 1

    print(f"Instrumentos: {dict(instr_counts)}")

    # --- Unidade amostral ---
    unidade_counts: Counter[str] = Counter()
    for r in aprovados:
        u = str(r.get("S2_unidade_espacial", "") or "").strip()
        if u:
            unidade_counts[normalize_unidade(u)] += 1

    print(f"Unidades amostrais: {dict(unidade_counts)}")

    # --- Metodos econometricos ---
    metodo_counts: Counter[str] = Counter()
    for r in aprovados:
        raw = str(r.get("S2_metodo_econometrico", "") or "").strip()
        methods = extract_metodos(raw)
        for m in methods:
            metodo_counts[m] += 1

    metodos_sorted = metodo_counts.most_common()
    print(f"Metodos: {len(metodos_sorted)} distintos")
    for m, c in metodos_sorted:
        print(f"  [{c}] {m}")

    # --- Autores (from RIS, with normalization) ---
    autor_counts = load_autores_from_ris()
    autores_sorted = autor_counts.most_common()
    print(f"Autores: {len(autores_sorted)} distintos (from RIS)")

    # --- IC ---
    ic_data = []
    for r in aprovados:
        pdf = str(r.get("Arquivo PDF", "") or "").replace(".pdf", "")
        ano = int(r.get("Ano", 0) or 0)
        pub = str(r.get("Publicado?", "") or "").strip() or "Nao"
        ic_val = float(r.get("IC") or 0)
        cit_pub = int(r.get("Citacoes (pub.)") or 0)
        cit_total = int(r.get("Citacoes (total)") or 0)

        # Determine type
        if pub == "Sim":
            tipo = "artigo publicado"
        else:
            base = str(r.get("Base", "") or "").lower()
            if "econpapers" in base:
                tipo = "texto para discussao"
            elif "anpec" in base:
                tipo = "apresentacao congresso"
            else:
                tipo = "nao-publicado (manual)"

        ic_data.append({
            "estudo": pdf, "ano": ano, "pub": pub,
            "cit": cit_pub, "n": cit_total,
            "ic": ic_val, "tipo": tipo,
        })

    ic_data.sort(key=lambda x: (-x["ic"], x["ano"]))
    print(f"IC entries: {len(ic_data)}")

    # ===================================================================
    # BUILD OUTPUT WORKBOOK
    # ===================================================================
    print(f"\nWriting {OUTPUT_PATH}...")
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Resumo"
    cur = 1

    # --- Visao Geral ---
    write_header(ws, cur, 1, "Visao Geral"); cur += 1
    write_col_header(ws, cur, 1, "Metrica")
    write_col_header(ws, cur, 2, "Valor"); cur += 1
    ws.cell(row=cur, column=1, value="Total de registros (pos-dedup)")
    ws.cell(row=cur, column=2, value=total); cur += 1
    ws.cell(row=cur, column=1, value="Aprovados")
    ws.cell(row=cur, column=2, value=n_aprov); cur += 1
    ws.cell(row=cur, column=1, value="Rejeitados")
    ws.cell(row=cur, column=2, value=n_rej); cur += 2

    # --- Motivos de Exclusao ---
    write_header(ws, cur, 1, "Motivos de Exclusao"); cur += 1
    write_col_header(ws, cur, 1, "Motivo")
    write_col_header(ws, cur, 2, "Qtd."); cur += 1
    for motivo in MOTIVO_ORDER:
        count = motivos.get(motivo, 0)
        if count > 0:
            ws.cell(row=cur, column=1, value=motivo)
            ws.cell(row=cur, column=2, value=count); cur += 1
    for motivo, count in motivos.items():
        if motivo not in MOTIVO_ORDER and count > 0:
            ws.cell(row=cur, column=1, value=motivo)
            ws.cell(row=cur, column=2, value=count); cur += 1
    cur += 1

    # --- Aprovados por Periodo ---
    write_header(ws, cur, 1, "Aprovados por Periodo (tab:estudos-ano)"); cur += 1
    write_col_header(ws, cur, 1, "Periodo")
    write_col_header(ws, cur, 2, "Qtd. Artigos"); cur += 1
    for label, _, _ in PERIODO_BINS:
        ws.cell(row=cur, column=1, value=label)
        ws.cell(row=cur, column=2, value=periodo_counts[label]); cur += 1
    c1 = ws.cell(row=cur, column=1, value="TOTAL"); c1.font = BOLD
    c2 = ws.cell(row=cur, column=2, value=n_aprov); c2.font = BOLD; cur += 2

    # --- Instrumentos PNDR ---
    write_header(ws, cur, 1, "Instrumentos PNDR (tab:instrumentos)"); cur += 1
    write_col_header(ws, cur, 1, "Instrumento")
    write_col_header(ws, cur, 2, "Qtd. Artigos"); cur += 1
    for instr in INSTRUMENTO_ORDER:
        ws.cell(row=cur, column=1, value=instr)
        ws.cell(row=cur, column=2, value=instr_counts.get(instr, 0)); cur += 1
    cur += 1

    # --- Unidade Amostral ---
    write_header(ws, cur, 1, "Unidade Amostral (tab:unidade-amostral)"); cur += 1
    write_col_header(ws, cur, 1, "Unidade Amostral")
    write_col_header(ws, cur, 2, "Qtd. Artigos"); cur += 1
    for u in UNIDADE_ORDER:
        count = unidade_counts.get(u, 0)
        if count > 0:
            ws.cell(row=cur, column=1, value=u)
            ws.cell(row=cur, column=2, value=count); cur += 1
    for u, count in unidade_counts.items():
        if u not in UNIDADE_ORDER and count > 0:
            ws.cell(row=cur, column=1, value=u)
            ws.cell(row=cur, column=2, value=count); cur += 1
    cur += 1

    # --- Metodos Econometricos ---
    write_header(ws, cur, 1, "Metodos Econometricos (tab:metodos)"); cur += 1
    write_col_header(ws, cur, 1, "Metodo")
    write_col_header(ws, cur, 2, "Estudos")
    write_col_header(ws, cur, 3, "MSM"); cur += 1
    for metodo, count in metodos_sorted:
        ws.cell(row=cur, column=1, value=metodo)
        ws.cell(row=cur, column=2, value=count)
        ws.cell(row=cur, column=3, value=MSM_SCALE.get(metodo, "n.c.")); cur += 1
    cur += 1

    # --- Autores ---
    write_header(ws, cur, 1, "Autores (tab:autores)"); cur += 1
    write_col_header(ws, cur, 1, "Autor")
    write_col_header(ws, cur, 2, "Qtd. Artigos"); cur += 1
    for autor, count in autores_sorted:
        ws.cell(row=cur, column=1, value=autor)
        ws.cell(row=cur, column=2, value=count); cur += 1
    cur += 1

    # --- Indice de Citacao Cruzada (IC) ---
    write_header(ws, cur, 1, "Indice de Citacao Cruzada (IC)"); cur += 1
    for col_n, label in enumerate(["Estudo", "Ano", "Pub?", "Cit.", "N", "IC", "Tipo"], 1):
        write_col_header(ws, cur, col_n, label)
    cur += 1
    for e in ic_data:
        ws.cell(row=cur, column=1, value=e["estudo"])
        ws.cell(row=cur, column=2, value=e["ano"])
        ws.cell(row=cur, column=3, value=e["pub"])
        ws.cell(row=cur, column=4, value=e["cit"])
        ws.cell(row=cur, column=5, value=e["n"])
        ws.cell(row=cur, column=6, value=e["ic"])
        ws.cell(row=cur, column=7, value=e["tipo"]); cur += 1

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 25

    wb_out.save(OUTPUT_PATH)
    print(f"Saved to {OUTPUT_PATH}")
    print("Done!")


if __name__ == "__main__":
    main()
